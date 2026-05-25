import datetime as dt
import io
from types import SimpleNamespace

from openpyxl import load_workbook

from web_comparativas.pliegos_fusion import (
    FUSION_PROCESS_FIELDS,
    calcular_estado_fusion,
    export_fusion_excel_bytes,
)


def _ns(**kwargs):
    return SimpleNamespace(**kwargs)


def _caso_base(garantias=None):
    datos_proceso = {
        "numero_proceso": "36/2026",
        "nombre_proceso": "Adquisición de sevoflurano",
        "organismo_contratante": "Gerencia de Contrataciones y Suministros",
        "objeto_contratacion": "Adquisición de sevoflurano",
        "tipo_proceso": "Licitación Pública",
        "fecha_apertura": "20/05/2026",
        "tipo_cotizacion": "Por Renglón",
        "tipo_adjudicacion": "Por Renglón",
        "plazo_mantenimiento_oferta": "60 días corridos",
        "moneda": "Moneda nacional",
        "duracion_contrato": "6 meses",
        "expediente": "648/2026",
        "modalidad": "Sin modalidad",
        "lleva_muestras": "Si",
    }
    return _ns(
        id=1,
        numero_proceso="52/25",
        fusion_cabecera=None,
        datos_proceso=_ns(datos=datos_proceso),
        control_carga=None,
        trazabilidad=[],
        analitica=None,
        hallazgos=[],
        cronograma=[
            _ns(hito="Acto de apertura", fecha="20/05/2026", hora="11:00"),
            _ns(hito="Presentación de ofertas", fecha="20/05/2026", hora="11:00"),
        ],
        garantias=garantias or [],
        requisitos=[],
        fusion_renglones=[
            _ns(
                numero_renglon="1",
                codigo_item="123",
                descripcion="Sevoflurano",
                cantidad="10",
                unidad="unidad",
                precio_unitario_estimado="",
                datos_extra={"obj_gasto": "2.5.2"},
            )
        ],
        renglones=[],
        faltantes=[],
    )


def _excel_proc_row(caso):
    payload = export_fusion_excel_bytes(caso)
    wb = load_workbook(io.BytesIO(payload))
    return wb["Datos del proceso"]


def _excel_workbook(caso):
    payload = export_fusion_excel_bytes(caso)
    return load_workbook(io.BytesIO(payload))


def test_fusion_process_field_count_is_32():
    ctx = calcular_estado_fusion(_caso_base())
    assert len(FUSION_PROCESS_FIELDS) == 32
    assert ctx["campos_total_obligatorios"] == 32


def test_uses_manual_request_id_for_id_proceso():
    caso = _caso_base()
    ctx = calcular_estado_fusion(caso)
    wb = _excel_workbook(caso)

    assert ctx["campos_fusion"]["id_proceso"]["valor"] == "52/25"
    assert wb["Datos del proceso"]["A2"].value == "52/25"
    assert wb["Datos del proceso"]["B2"].value == "36/2026"


def test_detail_obj_gas_and_cod_item_are_always_empty():
    wb = _excel_workbook(_caso_base())
    ws = wb["Detalle"]

    assert ws["B2"].value is None
    assert ws["C2"].value is None


def test_normalizes_currency_and_row_catalogs():
    ws = _excel_proc_row(_caso_base())
    assert ws["G2"].value == "Por renglones: parcial"
    assert ws["H2"].value == "Por renglones: parcial"
    assert ws["R2"].value == "ARS - Peso Argentino"


def test_preserves_opening_datetime_from_cronograma():
    ws = _excel_proc_row(_caso_base())
    assert ws["P2"].value == dt.datetime(2026, 5, 20, 11, 0)
    assert ws["P2"].number_format == "dd/mm/yyyy\\ hh:mm:ss"


def test_percentage_guarantees_are_text_or_empty():
    caso = _caso_base(
        [
            _ns(tipo="Garantía de Mantenimiento de Oferta", requerida="Si", porcentaje="5", base_calculo="", plazo="", formas_admitidas=""),
            _ns(tipo="Garantía de Anticipo Financiero", requerida="No", porcentaje="", base_calculo="", plazo="", formas_admitidas=""),
            _ns(tipo="Garantía de Cumplimiento de Contrato", requerida="Si", porcentaje="2.5", base_calculo="", plazo="", formas_admitidas=""),
        ]
    )
    ws = _excel_proc_row(caso)
    assert ws["X2"].value == "5%"
    assert ws["X2"].number_format == "@"
    assert ws["Y2"].value is None
    assert ws["Y2"].number_format == "@"
    assert ws["Z2"].value == "2,5%"
    assert ws["Z2"].number_format == "@"


def test_contragarantia_defaults_to_no_for_fusion():
    ws = _excel_proc_row(_caso_base())
    assert ws["AC2"].value == "No"
