"""Importa propuestas de Match (Mercado Privado) desde un Excel ya calculado.

FASE 1: NO se carga el dataset grande (~1,5M filas). Este importer lee SOLO la hoja
"Todos" (niveles A–D, ~65.990 filas accionables; el nivel E no tiene candidato y queda
fuera de la cola) en modo streaming con openpyxl read_only, y la vuelca a la tabla
compacta `match_propuestas`.

Patrón run-scoped (igual que Dimensionamiento): cada corrida crea una fila nueva en
`match_import_runs` en estado `pending_approval` y NO pisa la corrida vigente hasta
aprobar. La capa de servicio sirve SIEMPRE la última corrida `approved`.

APRENDIZAJE PREVIO (Dimensionamiento): la ruta del archivo se recibe SIEMPRE por
parámetro `--xlsx-path`. No hay default hardcodeado (un default roto trae problemas).

Uso local (desde la raíz del proyecto, con el venv activado):
    python scripts/import_match_propuestas.py --xlsx-path "ruta/al/propuestas.xlsx"
    python scripts/import_match_propuestas.py --xlsx-path "ruta.xlsx" --approve   # marca vigente

Opera sobre la base configurada por la app (local: web_comparativas/app.db).
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from collections import Counter
from pathlib import Path

# Permitir ejecución como script suelto agregando la raíz al path.
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl  # noqa: E402
from sqlalchemy.dialects.postgresql import insert as pg_insert  # noqa: E402
from sqlalchemy.dialects.sqlite import insert as sqlite_insert  # noqa: E402

from web_comparativas.match.models import (  # noqa: E402
    MATCH_RUN_APPROVED,
    MATCH_RUN_FAILED,
    MATCH_RUN_PENDING,
    MatchHomologacion,
    MatchHomologacionEvento,
    MatchImportRun,
    MatchPropuesta,
)
from web_comparativas.models import Base, IS_SQLITE, SessionLocal, engine  # noqa: E402


def _ensure_tables() -> None:
    """Crea las tablas match_* si faltan (idempotente). Permite correr el importer en
    una base donde la app aún no levantó tras agregar el módulo."""
    Base.metadata.create_all(
        bind=engine,
        tables=[
            MatchImportRun.__table__,
            MatchPropuesta.__table__,
            MatchHomologacion.__table__,
            MatchHomologacionEvento.__table__,
        ],
    )

SHEET_NAME = "Todos"
BATCH_SIZE = 5000

# Mapeo: columna destino (modelo) <- header en el Excel (hoja "Todos").
COLUMN_MAP = {
    "producto_plataforma": "producto_plataforma",
    "nivel_confianza": "nivel_confianza",
    "score_mejor": "score_mejor",
    "candidato_codigo": "candidato_1_codigo",
    "candidato_descripcion": "candidato_1_descripcion",
    "score_tfidf": "candidato_1_score_tfidf",
    "score_fuzzy": "candidato_1_score_fuzzy",
    "score_pharma": "candidato_1_score_pharma",
}


def _to_code(value) -> str | None:
    """Convierte el código del candidato a TEXT, sin '.0' espurio de floats de Excel."""
    if value is None:
        return None
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    return s or None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_nivel(value) -> str | None:
    s = _to_text(value)
    return s.upper()[:2] if s else None


def _resolve_headers(header_row: tuple) -> dict[str, int]:
    """Mapea cada header esperado a su índice de columna. Falla claro si falta alguno."""
    norm = {
        (str(c).strip().lower() if c is not None else ""): i
        for i, c in enumerate(header_row)
    }
    idx: dict[str, int] = {}
    faltantes: list[str] = []
    for dest, src in COLUMN_MAP.items():
        pos = norm.get(src.lower())
        if pos is None:
            faltantes.append(src)
        else:
            idx[dest] = pos
    if faltantes:
        raise SystemExit(
            f"[MATCH][IMPORT] Faltan columnas en la hoja '{SHEET_NAME}': {faltantes}. "
            f"Headers encontrados: {[c for c in header_row]}"
        )
    return idx


def _insert_batch(session, batch: list[dict]) -> None:
    """Inserta un lote idempotente (on_conflict_do_nothing por el UNIQUE de la corrida)."""
    if not batch:
        return
    if IS_SQLITE:
        stmt = sqlite_insert(MatchPropuesta).on_conflict_do_nothing()
    else:
        stmt = pg_insert(MatchPropuesta).on_conflict_do_nothing(
            constraint="uq_match_propuestas_run_prod_cand"
        )
    session.execute(stmt, batch)
    session.commit()


def importar(xlsx_path: Path, approve: bool) -> dict:
    if not xlsx_path.exists():
        raise SystemExit(f"[MATCH][IMPORT] Archivo no encontrado: {xlsx_path}")

    _ensure_tables()
    session = SessionLocal()
    run = MatchImportRun(
        source_path=str(xlsx_path),
        status=MATCH_RUN_PENDING,
        started_at=dt.datetime.utcnow(),
    )
    session.add(run)
    session.commit()
    session.refresh(run)
    run_id = run.id
    print(f"[MATCH][IMPORT] Corrida {run_id} iniciada (status={MATCH_RUN_PENDING}) desde {xlsx_path}")

    rows_inserted = 0
    codigos: set[str] = set()
    niveles: Counter = Counter()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise SystemExit(
                f"[MATCH][IMPORT] La hoja '{SHEET_NAME}' no existe. Hojas: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise SystemExit(f"[MATCH][IMPORT] La hoja '{SHEET_NAME}' está vacía.")
        idx = _resolve_headers(header)

        now = dt.datetime.utcnow()
        batch: list[dict] = []
        for raw in rows_iter:
            if raw is None:
                continue
            producto = _to_text(raw[idx["producto_plataforma"]])
            if not producto:
                continue  # sin descripción de portal no hay propuesta accionable
            codigo = _to_code(raw[idx["candidato_codigo"]])
            nivel = _to_nivel(raw[idx["nivel_confianza"]])

            batch.append({
                "import_run_id": run_id,
                "producto_plataforma": producto,
                "nivel_confianza": nivel,
                "score_mejor": _to_float(raw[idx["score_mejor"]]),
                "candidato_codigo": codigo,
                "candidato_descripcion": _to_text(raw[idx["candidato_descripcion"]]),
                "score_tfidf": _to_float(raw[idx["score_tfidf"]]),
                "score_fuzzy": _to_float(raw[idx["score_fuzzy"]]),
                "score_pharma": _to_float(raw[idx["score_pharma"]]),
                "created_at": now,
            })
            rows_inserted += 1
            if codigo:
                codigos.add(codigo)
            if nivel:
                niveles[nivel] += 1

            if len(batch) >= BATCH_SIZE:
                _insert_batch(session, batch)
                batch.clear()
                print(f"[MATCH][IMPORT]   ... {rows_inserted:,} filas procesadas", flush=True)

        _insert_batch(session, batch)
        wb.close()

        run.rows_inserted = rows_inserted
        run.articulos_distintos = len(codigos)
        run.counts_by_nivel = dict(sorted(niveles.items()))
        run.finished_at = dt.datetime.utcnow()
        if approve:
            run.status = MATCH_RUN_APPROVED
            run.approved_at = dt.datetime.utcnow()
            run.approved_by = "cli:import_match_propuestas"
        session.commit()
    except SystemExit:
        run.status = MATCH_RUN_FAILED
        run.error_message = "Abortado durante el import (ver consola)."
        run.finished_at = dt.datetime.utcnow()
        session.commit()
        raise
    except Exception as exc:
        session.rollback()
        run = session.get(MatchImportRun, run_id)
        if run is not None:
            run.status = MATCH_RUN_FAILED
            run.error_message = str(exc)[:1000]
            run.finished_at = dt.datetime.utcnow()
            session.commit()
        raise
    finally:
        session.close()

    return {
        "run_id": run_id,
        "status": MATCH_RUN_APPROVED if approve else MATCH_RUN_PENDING,
        "rows_inserted": rows_inserted,
        "articulos_distintos": len(codigos),
        "counts_by_nivel": dict(sorted(niveles.items())),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa propuestas de Match (hoja 'Todos').")
    parser.add_argument(
        "--xlsx-path",
        required=True,
        help="Ruta al Excel de propuestas (OBLIGATORIO, sin default).",
    )
    parser.add_argument(
        "--approve",
        action="store_true",
        help="Marca la corrida como 'approved' (vigente) al terminar. "
             "Sin esta bandera, queda 'pending_approval' (no se sirve hasta aprobar).",
    )
    args = parser.parse_args()

    result = importar(Path(args.xlsx_path).expanduser().resolve(), approve=args.approve)

    print("\n[MATCH][IMPORT] === Resumen de la corrida ===")
    print(f"  Corrida:             {result['run_id']}  (status={result['status']})")
    print(f"  Filas insertadas:    {result['rows_inserted']:,}")
    print(f"  Articulos distintos: {result['articulos_distintos']:,}  (candidato_codigo)")
    print(f"  Conteo por nivel:    {result['counts_by_nivel']}")
    total = result["rows_inserted"]
    print(f"  Conteo de control:   total={total:,}  (esperado ~= 65.990)")
    if result["status"] == MATCH_RUN_PENDING:
        print("  NOTA: corrida en 'pending_approval' (no visible aun). "
              "Reejecuta con --approve o aproba desde el flujo correspondiente.")
    print("[MATCH][IMPORT] ============================")


if __name__ == "__main__":
    main()
