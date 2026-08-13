"""Auditoría reproducible del puente nominal usado para cuentas CRM.

Solo lectura sobre los tres maestros. Escribe un CSV UTF-8 con BOM para revisión.
No consulta ni crea Opportunities en CRM.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from sqlalchemy import select

from web_comparativas.dimensionamiento.account_resolution import (
    MAX_CANDIDATES, normalize_cuit, normalize_identifier,
)
from web_comparativas.dimensionamiento.identity import canon
from web_comparativas.dimensionamiento.models import OportunidadSummary
from web_comparativas.models import SessionLocal

GENERIC_TOKENS = {
    "HOSPITAL", "CLINICA", "SANATORIO", "INSTITUTO", "CENTRO", "FUNDACION",
    "SOCIEDAD", "ASOCIACION", "MEDICA", "MEDICO", "SALUD", "SERVICIOS",
    "PRIVADO", "PRIVADA", "SA", "SRL", "SAS", "S", "A",
}


def generic_key(key: str) -> bool:
    return not [token for token in key.split() if token not in GENERIC_TOKENS and len(token) > 2]


def join(values) -> str:
    return " | ".join(sorted(str(value) for value in values if value not in (None, "")))


def audit(root: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    data = root / "web_comparativas" / "data"
    dataset = data / "archivos dimensionamiento" / "dataset_unificado_valorizado_2025_2026.csv"
    clients = data / "forecast_data" / "clientes.csv"
    operators_path = data / "archivos dimensionamiento" / "Operadores.xlsx"

    name_cuits: dict[str, set[str]] = defaultdict(set)
    name_homologated: dict[str, set[str]] = defaultdict(set)
    name_original: dict[str, set[str]] = defaultdict(set)
    name_accounts: dict[str, set[str]] = defaultdict(set)
    cuit_names: dict[str, set[str]] = defaultdict(set)
    cuit_accounts: dict[str, set[str]] = defaultdict(set)
    with dataset.open(encoding="utf-8-sig", newline="") as source:
        for row in csv.DictReader(source, delimiter=";"):
            cuit = normalize_cuit(row.get("cuit"))
            homologated = (row.get("cliente_nombre_homologado") or "").strip()
            key = canon(homologated)
            if not cuit or not key or key == "SIN DATO":
                continue
            name_cuits[key].add(cuit)
            name_homologated[key].add(homologated)
            name_original[key].add((row.get("cliente_nombre_original") or "").strip())
            cuit_names[cuit].add(key)
            account = normalize_identifier(row.get("cuenta_interna"))
            if account:
                name_accounts[key].add(account)
                cuit_accounts[cuit].add(account)

    client_codes: dict[str, set[str]] = defaultdict(set)
    client_names: dict[str, set[str]] = defaultdict(set)
    code_names: dict[str, set[str]] = defaultdict(set)
    with clients.open(encoding="utf-8-sig", newline="") as source:
        for row in csv.DictReader(source):
            key = canon(row.get("nombre"))
            code = normalize_identifier(row.get("codigo"))
            if not key or not code:
                continue
            client_codes[key].add(code)
            client_names[key].add((row.get("nombre") or "").strip())
            code_names[code].add(key)

    operators: dict[str, str] = {}
    workbook = load_workbook(operators_path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values)]
    for values_row in values:
        row = dict(zip(headers, values_row))
        code = normalize_identifier(row.get("Codigo"))
        if code:
            vendor = normalize_identifier(row.get("Vendedor"))
            name = str(row.get("Nombre") or "").strip()
            operators[code] = " - ".join(value for value in (vendor, name) if value)
    workbook.close()

    opportunity_counts: dict[str, int] = defaultdict(int)
    opportunity_total = 0
    opportunity_ambiguous = 0
    opportunity_without_relation = 0
    opportunity_with_alternatives = 0
    opportunity_only_original = 0
    try:
        with SessionLocal() as db:
            run_id = db.scalar(
                select(OportunidadSummary.import_run_id)
                .order_by(OportunidadSummary.import_run_id.desc()).limit(1)
            )
            opportunities = db.execute(
                select(OportunidadSummary).where(OportunidadSummary.import_run_id == run_id)
            ).scalars().all()
            opportunity_total = len(opportunities)
            for opportunity in opportunities:
                cuit = normalize_cuit(opportunity.cuit)
                keys = cuit_names.get(cuit or "", set())
                for key in keys:
                    opportunity_counts[key] += 1
                reasons = []
                if not cuit or not keys:
                    opportunity_without_relation += 1
                    continue
                if len(keys) > 1:
                    reasons.append("CUIT_CON_MULTIPLES_NOMBRES_HOMOLOGADOS")
                codes: set[str] = set()
                for key in keys:
                    codes.update(client_codes.get(key, set()))
                    if len(name_cuits[key]) > 1:
                        reasons.append("NOMBRE_CANONICO_COMPARTIDO_POR_MULTIPLES_CUIT")
                    if generic_key(key):
                        reasons.append("CLAVE_NOMINAL_SIN_TERMINOS_DISTINTIVOS")
                if len(codes | {normalize_identifier(opportunity.cuenta_interna)}) > MAX_CANDIDATES:
                    reasons.append("MAS_DE_25_CUENTAS_CANDIDATAS")
                if reasons:
                    opportunity_ambiguous += 1
                elif not codes:
                    opportunity_without_relation += 1
                elif codes - {normalize_identifier(opportunity.cuenta_interna)}:
                    opportunity_with_alternatives += 1
                else:
                    opportunity_only_original += 1
    except Exception:
        # El CSV de fuentes sigue siendo válido aunque no haya base local disponible.
        pass

    rows: list[dict[str, Any]] = []
    for key in sorted(name_cuits):
        reasons = []
        cuits = name_cuits[key]
        if len(cuits) > 1:
            reasons.append("NOMBRE_CANONICO_COMPARTIDO_POR_MULTIPLES_CUIT")
        if any(len(cuit_names[cuit]) > 1 for cuit in cuits):
            reasons.append("CUIT_CON_MULTIPLES_NOMBRES_HOMOLOGADOS")
        if generic_key(key):
            reasons.append("CLAVE_NOMINAL_SIN_TERMINOS_DISTINTIVOS")
        codes = client_codes.get(key, set())
        all_candidates = codes | {
            account for cuit in cuits for account in cuit_accounts.get(cuit, set())
        }
        if len(all_candidates) > MAX_CANDIDATES:
            reasons.append("MAS_DE_25_CUENTAS_CANDIDATAS")
        decision = "BLOQUEADA" if reasons else ("ACEPTADA_EXACTA_NO_AMBIGUA" if codes else "SIN_RELACION_CLIENTES")
        code_operators = [f"{code}: {operators[code]}" for code in sorted(codes) if operators.get(code)]
        rows.append({
            "cuit": join(cuits),
            "dataset_cliente_nombre_original": join(name_original[key]),
            "dataset_cliente_nombre_homologado": join(name_homologated[key]),
            "clientes_nombre_original": join(client_names.get(key, set())),
            "nombre_canonico": key,
            "cuentas_dataset": join(name_accounts.get(key, set())),
            "clientes_codigos_cuentas": join(codes),
            "operadores_por_cuenta": join(code_operators),
            "cantidad_cuit_asociados": len(cuits),
            "cantidad_nombres_del_cuit": max((len(cuit_names[cuit]) for cuit in cuits), default=0),
            "cantidad_cuentas_clientes": len(codes),
            "cantidad_candidatas_total": len(all_candidates),
            "cantidad_oportunidades": opportunity_counts.get(key, 0),
            "decision": decision,
            "motivo": join(reasons) if reasons else (
                "Igualdad exacta de canon(cliente_nombre_homologado) y canon(clientes.nombre); clave única para un CUIT."
            ),
        })
    rows.sort(key=lambda row: (
        row["decision"] != "BLOQUEADA",
        -row["cantidad_cuit_asociados"],
        -row["cantidad_nombres_del_cuit"],
        -row["cantidad_candidatas_total"],
        row["nombre_canonico"],
    ))
    for index, row in enumerate(rows, start=1):
        row["ranking_revision"] = index
        row["top_30_revision"] = "SI" if index <= 30 else "NO"

    matched = set(name_cuits) & set(client_codes)
    metrics = {
        "dataset_nombres_canonicos_unicos": len(name_cuits),
        "clientes_nombres_canonicos_unicos": len(client_codes),
        "coincidencias_exactas": len(matched),
        "coincidencias_uno_a_uno": sum(len(name_cuits[key]) == 1 and len(client_codes[key]) == 1 for key in matched),
        "coincidencias_uno_a_varios": sum(len(name_cuits[key]) == 1 and len(client_codes[key]) > 1 for key in matched),
        "coincidencias_varios_cuit_a_un_nombre": sum(len(value) > 1 for value in name_cuits.values()),
        "cuit_con_mas_de_un_nombre": sum(len(value) > 1 for value in cuit_names.values()),
        "nombres_con_mas_de_una_cuenta_dataset": sum(len(value) > 1 for value in name_accounts.values()),
        "nombres_con_mas_de_un_clientes_codigo": sum(len(client_codes[key]) > 1 for key in matched),
        "clientes_codigo_con_mas_de_un_nombre": sum(len(value) > 1 for value in code_names.values()),
        "claves_genericas_matcheadas": sum(generic_key(key) for key in matched),
        "maximo_candidatas_en_oportunidad": max((row["cantidad_candidatas_total"] for row in rows), default=0),
        "relaciones_mas_de_25_en_datos_actuales": sum(row["cantidad_candidatas_total"] > MAX_CANDIDATES for row in rows),
        "oportunidades_total": opportunity_total,
        "oportunidades_con_alternativas": opportunity_with_alternatives,
        "oportunidades_ambiguas": opportunity_ambiguous,
        "oportunidades_sin_relacion": opportunity_without_relation,
        "oportunidades_solo_cuenta_original": opportunity_only_original,
    }
    return metrics, rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--metrics-output", type=Path)
    args = parser.parse_args()
    metrics, rows = audit(args.root.resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "ranking_revision", "top_30_revision", "cuit",
        "dataset_cliente_nombre_original", "dataset_cliente_nombre_homologado",
        "clientes_nombre_original", "nombre_canonico", "cuentas_dataset",
        "clientes_codigos_cuentas", "operadores_por_cuenta", "cantidad_cuit_asociados",
        "cantidad_nombres_del_cuit", "cantidad_cuentas_clientes",
        "cantidad_candidatas_total", "cantidad_oportunidades", "decision", "motivo",
    ]
    with args.output.open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    if args.metrics_output:
        args.metrics_output.parent.mkdir(parents=True, exist_ok=True)
        args.metrics_output.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"metrics": metrics, "rows": len(rows), "output": str(args.output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()