"""
pliegos_fusion.py
Validación Fusion independiente del campo listo_para_fusion de LicIA.
Implementa la matriz de campos obligatorios de Verónica y calcula el estado_fusion
internamente, sin confiar en Control_Carga ni Fusion_Cabecera ciegamente.
"""
from __future__ import annotations

import io
import re
import unicodedata
import datetime as dt
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Marcadores de valor faltante — nunca deben tratarse como dato válido
# ---------------------------------------------------------------------------
FUSION_MISSING_MARKERS = {
    "", "-", "--", "n/a", "na", "nan", "none",
    "no encontrado", "no encontrada", "no identificado", "no identificada",
    "sin dato", "sin datos", "sin dato en pliego", "sin informacion", "sin información",
    "pendiente", "a definir", "a confirmar", "por definir", "por confirmar",
    "no aplica", "no disponible", "no informado",
    "pendiente validacion", "pendiente validación", "faltante fusion", "faltante fusión",
    "ambiguo", "contradictorio", "vacio en excel fusion", "vacío en excel fusión",
}

BASE_DIR = Path(__file__).resolve().parent
FUSION_TEMPLATE_PATH = BASE_DIR / "templates_excel" / "Documento_Importacion_Fusion_template.xlsx"


def _norm(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    return re.sub(r"\s+", " ", "".join(ch for ch in nfkd if not unicodedata.combining(ch))).strip().lower()


def _safe(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in {"nan", "none"} else s


def is_fusion_missing(value) -> bool:
    """True cuando el valor debe considerarse faltante para Fusion."""
    return _norm(value) in FUSION_MISSING_MARKERS


def _pick(*values) -> str:
    """Devuelve el primer valor no-faltante de la lista de candidatos."""
    for v in values:
        s = _safe(v)
        if s and not is_fusion_missing(s):
            return s
    return ""


def get_manual_id_proceso(caso) -> str:
    """ID Proceso informado por el usuario al crear la solicitud."""
    return _pick(getattr(caso, "numero_proceso", None))


def _build_licia_texto(estado_licia: dict) -> str:
    """Construye texto legible del estado informado por el análisis automático."""
    if not estado_licia or not isinstance(estado_licia, dict):
        return ""

    def _yn(val) -> str:
        n = _norm(val)
        if n in {"si", "sí", "yes", "true", "1"}:
            return "Sí"
        if n in {"no", "false", "0"}:
            return "No"
        return ""

    parts = []

    listo = _yn(estado_licia.get("listo_para_fusion", ""))
    if listo:
        parts.append(f"Listo: {listo}")

    nivel = _safe(estado_licia.get("nivel_riesgo", ""))
    if nivel:
        parts.append(f"Riesgo: {nivel.title()}")

    rev = _yn(estado_licia.get("requiere_revision_analista", ""))
    if rev == "Sí":
        parts.append("Requiere revisión")

    motivo = _safe(estado_licia.get("motivo_estado", ""))
    if motivo and len(motivo) < 60:
        parts.append(motivo)

    return " · ".join(parts) if parts else ""


# ---------------------------------------------------------------------------
# Catálogos de normalización para Fusion
# ---------------------------------------------------------------------------
PROCEDIMIENTO_CATALOGO = {
    "licitacion publica": "Licitación Pública",
    "licitacion privada": "Licitación Privada",
    "concurso publico": "Concurso Público",
    "concurso privado": "Concurso Privado",
    "contratacion directa": "Contratación Directa",
    "contratacion por cotejo": "Cotejo",
    "cotejo de precios": "Cotejo",
    "cotejo": "Cotejo",
    "licitacion": "Licitación Pública",
    "concurso": "Concurso Público",
}

MODALIDAD_CATALOGO = {
    "sin modalidad": "Sin Modalidad",
    "convenio marco": "Convenio Marco",
    "compra consolidada": "Compra Consolidada",
    "iniciativa privada": "Iniciativa Privada",
    "subasta inversa": "Subasta Inversa",
    "sin especificar": "Sin Especificar",
}

TIPO_COTIZACION_CATALOGO = {
    "total": "Total",
    "parcial": "Por renglones: parcial",
    "por renglon": "Por renglones: parcial",
    "por renglones": "Por renglones: parcial",
    "por renglón": "Por renglones: parcial",
    "por renglónes": "Por renglones: parcial",
    "global": "Global",
}

TIPO_ADJUDICACION_CATALOGO = {
    "total": "Total",
    "parcial": "Por renglones: parcial",
    "por renglon": "Por renglones: parcial",
    "por renglones": "Por renglones: parcial",
    "por renglón": "Por renglones: parcial",
    "por renglónes": "Por renglones: parcial",
}

MONEDA_CATALOGO = {
    "moneda nacional": "ARS - Peso Argentino",
    "pesos": "ARS - Peso Argentino",
    "peso argentino": "ARS - Peso Argentino",
    "ars": "ARS - Peso Argentino",
    "$": "ARS - Peso Argentino",
}


def _normalize_catalog(value: str, catalog: dict, field_key: str) -> dict:
    norm = _norm(value)
    # Buscar coincidencia exacta
    if norm in catalog:
        return {"valor_normalizado": catalog[norm], "requiere_validacion": False}
    # Buscar si el valor contiene alguna clave
    for key, mapped in catalog.items():
        if key in norm:
            return {"valor_normalizado": mapped, "requiere_validacion": False}
    # Sin coincidencia: marcar para revisión manual
    return {"valor_normalizado": value, "requiere_validacion": True}


def _normalize_bool_sino(value) -> tuple[str, bool]:
    text = _clean_export_value(value)
    if not text:
        return "", False
    norm = _norm(text)
    token = re.split(r"[\s,;:/()]+", norm)[0] if norm else ""
    if token in {"si", "s", "yes", "true", "1"}:
        return "Si", False
    if token in {"no", "false", "0"}:
        return "No", False
    return text, True


def _normalize_moneda(value) -> tuple[str, bool]:
    text = _clean_export_value(value)
    if not text:
        return "", False
    norm = _norm(text)
    if norm in MONEDA_CATALOGO:
        return MONEDA_CATALOGO[norm], False
    return text, True


# ---------------------------------------------------------------------------
# Resolución de campos Fusion con prioridad y fallback
# ---------------------------------------------------------------------------

def _get_fusion_cabecera(caso) -> dict:
    if caso.fusion_cabecera and caso.fusion_cabecera.datos:
        return {k.lower().strip(): v for k, v in caso.fusion_cabecera.datos.items()}
    return {}


def _get_proceso(caso) -> dict:
    if caso.datos_proceso and caso.datos_proceso.datos:
        return {k.lower().strip(): v for k, v in caso.datos_proceso.datos.items()}
    return {}


def _get_control_carga(caso) -> dict:
    if caso.control_carga and caso.control_carga.datos:
        datos = caso.control_carga.datos
        if isinstance(datos, list) and datos:
            return {k.lower().strip(): v for k, v in datos[0].items()}
        if isinstance(datos, dict):
            return {k.lower().strip(): v for k, v in datos.items()}
    return {}


def _from_proceso(proceso: dict, *keys) -> str:
    for key in keys:
        for k, v in proceso.items():
            if k == key.lower() or k.replace(" ", "_") == key.lower() or k.replace("_", " ") == key.lower():
                val = _safe(v)
                if val and not is_fusion_missing(val):
                    return val
    return ""


def _from_trazabilidad(caso, *campos) -> str:
    campos_norm = [_norm(c) for c in campos]
    for t in (caso.trazabilidad or []):
        campo_n = _norm(t.campo)
        if any(campo_n == cn or campo_n.startswith(cn) or cn.startswith(campo_n) for cn in campos_norm if cn):
            val = _safe(t.valor_extraido)
            if val and not is_fusion_missing(val):
                return val
    return ""


def _from_cronograma(caso, *palabras_clave) -> str:
    palabras = [_norm(p) for p in palabras_clave]
    for hito in (caso.cronograma or []):
        hito_norm = _norm(hito.hito)
        if any(p in hito_norm for p in palabras):
            fecha = _safe(hito.fecha)
            hora = _safe(hito.hora)
            if fecha and not is_fusion_missing(fecha):
                return f"{fecha} {hora}".strip() if hora and not is_fusion_missing(hora) else fecha
    return ""


def _fecha_apertura_con_hora(caso, proc: dict, fc: dict) -> tuple[str, str]:
    """Prioriza el valor base disponible y agrega hora desde cronograma si falta."""
    base = _pick(
        fc.get("fecha_apertura_fusion"),
        _from_proceso(proc, "fecha_apertura"),
        _from_trazabilidad(caso, "fecha_apertura"),
    )
    cronograma = _pick(
        _from_cronograma(caso, "acto de apertura", "presentacion de ofertas", "presentación de ofertas", "apertura"),
    )
    valor = base or cronograma
    fuente = "Fusion_Cabecera" if fc.get("fecha_apertura_fusion") and not is_fusion_missing(fc.get("fecha_apertura_fusion", "")) else (
        "Proceso" if base else "Cronograma"
    )
    if valor and cronograma:
        valor_dt = _parse_date_native(valor)
        cron_dt = _parse_date_native(cronograma)
        if valor_dt and cron_dt and not (valor_dt.hour or valor_dt.minute or valor_dt.second):
            if cron_dt.hour or cron_dt.minute or cron_dt.second:
                valor = dt.datetime(
                    valor_dt.year, valor_dt.month, valor_dt.day,
                    cron_dt.hour, cron_dt.minute, cron_dt.second,
                ).strftime("%d/%m/%Y %H:%M:%S")
                fuente = f"{fuente} + Cronograma"
    return valor, fuente


def _from_cronograma_campo(caso, *palabras_clave, campo: str = "fecha") -> str:
    palabras = [_norm(p) for p in palabras_clave]
    for hito in (caso.cronograma or []):
        hito_norm = _norm(hito.hito)
        if any(p in hito_norm for p in palabras):
            if campo == "hito":
                return _safe(hito.hito)
            elif campo == "fecha":
                val = _safe(hito.fecha)
                return val if val and not is_fusion_missing(val) else ""
    return ""


def _from_analitica(caso, *campos_relacionados) -> str:
    analitica = caso.analitica
    if not analitica:
        return ""
    datos = analitica.datos
    if not datos:
        return ""
    rows = datos if isinstance(datos, list) else []
    campos_norm = [_norm(c) for c in campos_relacionados]
    for row in rows:
        if not isinstance(row, dict):
            continue
        campo_rel = _norm(row.get("campo_relacionado", ""))
        if any(campo_rel == cn or cn in campo_rel for cn in campos_norm if cn):
            val = _safe(row.get("valor_analitico", ""))
            if val and not is_fusion_missing(val):
                return val
    return ""


def _from_hallazgos(caso, *campos_dashboard) -> str:
    campos_norm = [_norm(c) for c in campos_dashboard]
    for h in (caso.hallazgos or []):
        extra = getattr(h, "datos_extra", {}) or {}
        campo_dash = _norm(extra.get("campo_dashboard_sugerido", ""))
        if any(campo_dash == cn or cn in campo_dash for cn in campos_norm if cn):
            val = _safe(extra.get("valor_sugerido", ""))
            if val and not is_fusion_missing(val):
                return val
    return ""


# ---------------------------------------------------------------------------
# Resolvedor por campo Fusion
# ---------------------------------------------------------------------------

def _resolver_campo_fusion(caso, field_key: str, fc: dict, proc: dict) -> dict:
    """
    Resuelve un campo Fusion usando la cadena de prioridad definida por Verónica.
    Retorna dict con: valor, fuente, requiere_validacion, valor_normalizado
    """
    valor = ""
    fuente = ""
    valor_normalizado = ""
    requiere_validacion = False

    if field_key == "unidad_ejecutora":
        valor = _pick(
            fc.get("unidad_ejecutora_fusion"),
            _from_proceso(proc, "unidad_operativa"),
            _from_proceso(proc, "organismo_contratante"),
            _from_trazabilidad(caso, "unidad_operativa", "unidad_ejecutora"),
        )
        fuente = "Fusion_Cabecera" if fc.get("unidad_ejecutora_fusion") and not is_fusion_missing(fc["unidad_ejecutora_fusion"]) else "Proceso"

    elif field_key == "numero_proceso":
        valor = _pick(
            fc.get("numero_proceso_fusion"),
            _from_proceso(proc, "numero_proceso"),
            _from_trazabilidad(caso, "numero_proceso", "nro_proceso"),
        )
        fuente = "Fusion_Cabecera" if fc.get("numero_proceso_fusion") and not is_fusion_missing(fc.get("numero_proceso_fusion", "")) else "Proceso"

    elif field_key == "nombre_proceso":
        valor = _pick(
            fc.get("nombre_proceso_fusion"),
            _from_proceso(proc, "nombre_proceso"),
        )
        fuente = "Fusion_Cabecera" if fc.get("nombre_proceso_fusion") and not is_fusion_missing(fc.get("nombre_proceso_fusion", "")) else "Proceso"

    elif field_key == "procedimiento_seleccion":
        raw = _pick(
            _from_proceso(proc, "tipo_proceso"),
            _from_trazabilidad(caso, "procedimiento_seleccion", "tipo_proceso"),
        )
        if raw:
            norm_result = _normalize_catalog(raw, PROCEDIMIENTO_CATALOGO, field_key)
            valor = raw
            valor_normalizado = norm_result["valor_normalizado"]
            requiere_validacion = norm_result["requiere_validacion"]
            fuente = "Proceso"

    elif field_key == "objeto_contratacion":
        valor = _pick(
            fc.get("objeto_proceso_fusion"),
            _from_proceso(proc, "objeto_contratacion"),
            _from_trazabilidad(caso, "objeto_contratacion"),
        )
        fuente = "Fusion_Cabecera" if fc.get("objeto_proceso_fusion") and not is_fusion_missing(fc.get("objeto_proceso_fusion", "")) else "Proceso"

    elif field_key == "fecha_apertura":
        valor, fuente = _fecha_apertura_con_hora(caso, proc, fc)

    elif field_key == "tipo_cotizacion":
        raw = _pick(
            fc.get("tipo_cotizacion_fusion"),
            _from_proceso(proc, "tipo_cotizacion"),
            _from_trazabilidad(caso, "tipo_cotizacion"),
        )
        if raw:
            norm_result = _normalize_catalog(raw, TIPO_COTIZACION_CATALOGO, field_key)
            valor = raw
            valor_normalizado = norm_result["valor_normalizado"]
            requiere_validacion = norm_result["requiere_validacion"]
            fuente = "Fusion_Cabecera" if fc.get("tipo_cotizacion_fusion") and not is_fusion_missing(fc.get("tipo_cotizacion_fusion", "")) else "Proceso"

    elif field_key == "tipo_adjudicacion":
        raw = _pick(
            _from_proceso(proc, "tipo_adjudicacion"),
            _from_analitica(caso, "tipo_adjudicacion"),
            _from_hallazgos(caso, "tipo_adjudicacion"),
        )
        if raw:
            norm_result = _normalize_catalog(raw, TIPO_ADJUDICACION_CATALOGO, field_key)
            valor = raw
            valor_normalizado = norm_result["valor_normalizado"]
            requiere_validacion = norm_result["requiere_validacion"]
            fuente = "Proceso"

    elif field_key == "cant_oferta_permitidas":
        # Buscar con múltiples aliases
        proc_val = ""
        for alias in ["cant_oferta_permitidas", "cantidad_ofertas_permitidas", "acepta_mas_de_una_oferta",
                      "oferta_permitida", "cantidad_ofertas", "ofertas_permitidas", "cant ofertas", "ofertas"]:
            proc_val = _from_proceso(proc, alias)
            if proc_val:
                break
        if not proc_val:
            proc_val = _from_trazabilidad(caso, "cant_oferta_permitidas", "cantidad_ofertas_permitidas", "ofertas_permitidas")
        valor = proc_val
        fuente = "Proceso" if proc_val else ""

    elif field_key == "plazo_mantenimiento_oferta":
        valor = _pick(
            _from_proceso(proc, "plazo_mantenimiento_oferta"),
            _from_cronograma(caso, "mantenimiento de oferta", "mantenimiento oferta"),
            _from_trazabilidad(caso, "plazo_mantenimiento_oferta", "mantenimiento_oferta"),
        )
        fuente = "Proceso"

    elif field_key == "acepta_redeterminacion":
        raw = _pick(
            _from_proceso(proc, "acepta_redeterminacion"),
            _from_trazabilidad(caso, "acepta_redeterminacion"),
        )
        if raw:
            valor, requiere_validacion = _normalize_bool_sino(raw)
            fuente = "Proceso"

    elif field_key == "fecha_inicio_consulta":
        valor = _pick(
            _from_cronograma(caso, "inicio consulta", "inicio de consulta", "apertura consulta", "inicio consultas"),
            _from_proceso(proc, "fecha_inicio_consulta", "inicio_consultas", "inicio_consulta"),
            _from_trazabilidad(caso, "fecha_inicio_consulta", "inicio_consultas", "inicio_consulta"),
        )
        fuente = "Cronograma" if valor and _from_cronograma(caso, "inicio consulta", "inicio de consulta", "apertura consulta", "inicio consultas") else "Proceso"

    elif field_key == "fecha_final_consulta":
        valor = _pick(
            _from_cronograma(caso, "cierre consulta", "fin consulta", "final consulta", "limite consulta", "cierre de consulta", "fin de consulta"),
            _from_proceso(proc, "fecha_final_consulta", "cierre_consultas", "fin_consulta"),
            _from_trazabilidad(caso, "fecha_final_consulta", "cierre_consultas", "limite_consultas"),
        )
        fuente = "Cronograma" if valor and _from_cronograma(caso, "cierre consulta", "fin consulta", "final consulta", "limite consulta") else "Proceso"

    elif field_key == "monto":
        raw = _pick(
            _from_proceso(proc, "presupuesto_oficial"),
            _from_trazabilidad(caso, "presupuesto_oficial", "monto"),
            _from_analitica(caso, "presupuesto_oficial", "monto"),
        )
        valor = raw
        fuente = "Proceso"

    elif field_key == "moneda":
        raw = _pick(
            fc.get("moneda_fusion"),
            _from_proceso(proc, "moneda"),
            _from_trazabilidad(caso, "moneda"),
        )
        if raw:
            valor = raw
            valor_normalizado, requiere_validacion = _normalize_moneda(raw)
            fuente = "Fusion_Cabecera" if fc.get("moneda_fusion") and not is_fusion_missing(fc.get("moneda_fusion", "")) else "Proceso"

    elif field_key == "duracion_contrato":
        valor = _pick(
            _from_proceso(proc, "duracion_contrato"),
            _from_trazabilidad(caso, "duracion_contrato"),
        )
        fuente = "Proceso"

    elif field_key == "acepta_prorroga":
        raw = _pick(
            _from_proceso(proc, "acepta_prorroga"),
            _from_trazabilidad(caso, "acepta_prorroga"),
        )
        if raw:
            valor, requiere_validacion = _normalize_bool_sino(raw)
            fuente = "Proceso"

    elif field_key == "expediente":
        valor = _pick(
            fc.get("expediente_fusion"),
            _from_proceso(proc, "expediente"),
            _from_trazabilidad(caso, "expediente"),
        )
        fuente = "Fusion_Cabecera" if fc.get("expediente_fusion") and not is_fusion_missing(fc.get("expediente_fusion", "")) else "Proceso"

    elif field_key == "modalidad":
        raw = _pick(
            fc.get("modalidad_fusion"),
            _from_proceso(proc, "modalidad"),
            _from_trazabilidad(caso, "modalidad"),
        )
        if raw:
            norm_result = _normalize_catalog(raw, MODALIDAD_CATALOGO, field_key)
            valor = raw
            valor_normalizado = norm_result["valor_normalizado"]
            requiere_validacion = norm_result["requiere_validacion"]
            fuente = "Fusion_Cabecera" if fc.get("modalidad_fusion") and not is_fusion_missing(fc.get("modalidad_fusion", "")) else "Proceso"

    return {
        "valor": valor,
        "fuente": fuente,
        "valor_normalizado": valor_normalizado or valor,
        "requiere_validacion": requiere_validacion,
        "completo": bool(valor and not is_fusion_missing(valor)),
    }


# ---------------------------------------------------------------------------
# Fuente única de campos Fusion/ERP: hoja "Datos del proceso" A:AF
# ---------------------------------------------------------------------------
FUSION_PROCESS_FIELDS = [
    "ID Proceso",
    "N° Procesos",
    "Nombre Proceso",
    "Unidad Ejecutora",
    "Objeto Contratación",
    "Procedimiento Selección",
    "Tipo Cotización",
    "Tipo Adjudicación",
    "Cant. Oferta Permitidas",
    "Plazo Mantenimiento Oferta",
    "Acepta Redeterminación",
    "Ampliación",
    "Acepta Prórroga",
    "Fecha Inicio Consulta",
    "Fecha Final Consulta",
    "Fecha Acto Apertura",
    "Monto",
    "Moneda",
    "Duración Contrato",
    "Nro Expediente",
    "Modalidad",
    "Alternativa",
    "Lugar y Cond. De Entrega",
    "Garantía de Mantenimiento de Oferta",
    "Garantía de Anticipo Financiero",
    "Garantía de Cumplimiento de Contrato",
    "Garantía de Impugnación al Pliego",
    "Garantía de Impugnación a la Preadjudicación",
    "Garantía de Incorporar Contragarantía",
    "Exigencia con Causal de Desestimación",
    "¿Lleva Muestras?",
    "Fecha limite para presentación de muestras",
]

FUSION_PROCESS_FIELD_DEFS = [
    {"key": "id_proceso", "label": "ID Proceso", "categoria": "fusion_obligatorio"},
    {"key": "numero_proceso", "label": "N° Procesos", "categoria": "fusion_obligatorio"},
    {"key": "nombre_proceso", "label": "Nombre Proceso", "categoria": "fusion_obligatorio"},
    {"key": "unidad_ejecutora", "label": "Unidad Ejecutora", "categoria": "fusion_obligatorio"},
    {"key": "objeto_contratacion", "label": "Objeto Contratación", "categoria": "fusion_obligatorio"},
    {"key": "procedimiento_seleccion", "label": "Procedimiento Selección", "categoria": "fusion_obligatorio"},
    {"key": "tipo_cotizacion", "label": "Tipo Cotización", "categoria": "fusion_obligatorio"},
    {"key": "tipo_adjudicacion", "label": "Tipo Adjudicación", "categoria": "fusion_obligatorio"},
    {"key": "cant_oferta_permitidas", "label": "Cant. Oferta Permitidas", "categoria": "fusion_obligatorio"},
    {"key": "plazo_mantenimiento_oferta", "label": "Plazo Mantenimiento Oferta", "categoria": "fusion_obligatorio"},
    {"key": "acepta_redeterminacion", "label": "Acepta Redeterminación", "categoria": "fusion_obligatorio"},
    {"key": "ampliacion", "label": "Ampliación", "categoria": "fusion_obligatorio"},
    {"key": "acepta_prorroga", "label": "Acepta Prórroga", "categoria": "fusion_obligatorio"},
    {"key": "fecha_inicio_consulta", "label": "Fecha Inicio Consulta", "categoria": "fusion_obligatorio"},
    {"key": "fecha_final_consulta", "label": "Fecha Final Consulta", "categoria": "fusion_obligatorio"},
    {"key": "fecha_apertura", "label": "Fecha Acto Apertura", "categoria": "fusion_obligatorio"},
    {"key": "monto", "label": "Monto", "categoria": "fusion_obligatorio"},
    {"key": "moneda", "label": "Moneda", "categoria": "fusion_obligatorio"},
    {"key": "duracion_contrato", "label": "Duración Contrato", "categoria": "fusion_obligatorio"},
    {"key": "expediente", "label": "Nro Expediente", "categoria": "fusion_obligatorio"},
    {"key": "modalidad", "label": "Modalidad", "categoria": "fusion_obligatorio"},
    {"key": "alternativa", "label": "Alternativa", "categoria": "fusion_obligatorio"},
    {"key": "lugar_entrega", "label": "Lugar y Cond. De Entrega", "categoria": "fusion_obligatorio"},
    {"key": "garantia_mantenimiento_oferta", "label": "Garantía de Mantenimiento de Oferta", "categoria": "fusion_obligatorio"},
    {"key": "garantia_anticipo_financiero", "label": "Garantía de Anticipo Financiero", "categoria": "fusion_obligatorio"},
    {"key": "garantia_cumplimiento_contrato", "label": "Garantía de Cumplimiento de Contrato", "categoria": "fusion_obligatorio"},
    {"key": "garantia_impugnacion_pliego", "label": "Garantía de Impugnación al Pliego", "categoria": "fusion_obligatorio"},
    {"key": "garantia_impugnacion_preadjudicacion", "label": "Garantía de Impugnación a la Preadjudicación", "categoria": "fusion_obligatorio"},
    {"key": "garantia_contragarantia", "label": "Garantía de Incorporar Contragarantía", "categoria": "fusion_obligatorio"},
    {"key": "exigencia_desestimacion", "label": "Exigencia con Causal de Desestimación", "categoria": "fusion_obligatorio"},
    {"key": "lleva_muestras", "label": "¿Lleva Muestras?", "categoria": "fusion_obligatorio"},
    {"key": "fecha_limite_muestras", "label": "Fecha limite para presentación de muestras", "categoria": "fusion_obligatorio"},
]

FUSION_CAMPOS_OBLIGATORIOS = FUSION_PROCESS_FIELD_DEFS
FUSION_CAMPOS_COMPLEMENTARIOS = []

FUSION_CAMPOS_NO_OBLIGATORIOS = [
    {"key": "obj_gasto",                "label": "Obj. Gas.",                    "categoria": "no_obligatorio"},
    {"key": "codigo_item",              "label": "Cod. Item",                    "categoria": "no_obligatorio"},
]

FUSION_EXPORT_PROCESO_COLUMNS = FUSION_PROCESS_FIELDS

FUSION_EXPORT_RENGLON_COLUMNS = [
    "Item",
    "Obj. Gas.",
    "Cod. Item",
    "Descripción",
    "Cant",
]

_FUSION_PROCESS_FIELD_BY_COLUMN = {
    # Columna A: ID Proceso se resuelve desde el valor manual cargado por el usuario.
    "N° Procesos":                   "numero_proceso",
    "Nombre Proceso":                "nombre_proceso",
    "Unidad Ejecutora":              "unidad_ejecutora",
    "Objeto Contratación":           "objeto_contratacion",
    "Procedimiento Selección":       "procedimiento_seleccion",
    "Tipo Cotización":               "tipo_cotizacion",
    "Tipo Adjudicación":             "tipo_adjudicacion",
    "Cant. Oferta Permitidas":       "cant_oferta_permitidas",
    "Plazo Mantenimiento Oferta":    "plazo_mantenimiento_oferta",
    "Acepta Redeterminación":        "acepta_redeterminacion",
    "Acepta Prórroga":               "acepta_prorroga",
    "Fecha Inicio Consulta":         "fecha_inicio_consulta",
    "Fecha Final Consulta":          "fecha_final_consulta",
    "Fecha Acto Apertura":           "fecha_apertura",
    "Monto":                         "monto",
    "Moneda":                        "moneda",
    "Duración Contrato":             "duracion_contrato",
    "Nro Expediente":                "expediente",
    "Modalidad":                     "modalidad",
}

_FUSION_CORE_FIELD_KEYS = set(_FUSION_PROCESS_FIELD_BY_COLUMN.values())


# ---------------------------------------------------------------------------
# Cálculo de renglones Fusion
# ---------------------------------------------------------------------------

def _resolver_renglones_fusion(caso) -> list[dict]:
    """
    Prioriza Fusion_Renglones; enriquece con datos de Renglones cuando falte fuente/observaciones.
    Nunca devuelve lista vacía si hay renglones en alguna hoja.
    """
    renglones_out = []

    # Índice de renglones por orden para enriquecimiento
    renglones_idx: dict[int, object] = {}
    for r in (caso.renglones or []):
        renglones_idx[r.orden] = r

    fusion_renglones = caso.fusion_renglones or []
    renglones_base = caso.renglones or []

    # Preferir Fusion_Renglones si existen
    fuente_principal = fusion_renglones if fusion_renglones else renglones_base

    for idx, fr in enumerate(fuente_principal):
        if fusion_renglones:
            orden_str = _safe(fr.numero_renglon) or str(idx + 1)
            try:
                orden_int = int(float(orden_str))
            except (ValueError, TypeError):
                orden_int = idx + 1

            # Enriquecer con hoja Renglones si existe
            renglon_base = renglones_idx.get(orden_int)
            extra = fr.datos_extra or {}

            row = {
                "item": orden_str,
                "obj_gasto": _safe(extra.get("obj_gasto", "")) or _safe(extra.get("objeto_gasto", "")) or _safe(extra.get("obj_gas", "")),
                "codigo_item": _safe(fr.codigo_item),
                "descripcion": _safe(fr.descripcion),
                "cantidad": _safe(fr.cantidad),
                "unidad": _safe(fr.unidad),
                "precio_unitario": _safe(fr.precio_unitario_estimado) or _safe(extra.get("precio_unitario", "")),
                "importe_total": _safe(extra.get("importe_total", "")),
                "lugar_entrega": _safe(extra.get("lugar_entrega", "")),
                "plazo_entrega": _safe(extra.get("plazo_entrega", "")),
                "periodicidad": _safe(extra.get("periodicidad", "")),
                "marca": _safe(extra.get("marca", "")),
                "destino_efector": _safe(extra.get("destino_efector", "")),
                "observaciones": _safe(extra.get("observaciones", "")),
                "fuente": _safe(extra.get("fuente", "Fusion_Renglones")),
                "estado": _safe(extra.get("estado", "")),
            }
            # Enriquecer fuente/observaciones desde hoja Renglones
            if renglon_base:
                rb_extra = renglon_base.datos_extra or {}
                if not row["fuente"]:
                    row["fuente"] = _safe(rb_extra.get("fuente", "Renglones"))
                if not row["observaciones"]:
                    row["observaciones"] = _safe(renglon_base.obs_tecnicas)
                if not row["lugar_entrega"]:
                    row["lugar_entrega"] = _safe(rb_extra.get("lugar_entrega", ""))
                if not row["plazo_entrega"]:
                    row["plazo_entrega"] = _safe(rb_extra.get("plazo_entrega", ""))
                if not row["periodicidad"]:
                    row["periodicidad"] = _safe(rb_extra.get("periodicidad", ""))
                if not row["destino_efector"]:
                    row["destino_efector"] = _safe(renglon_base.destino_efector)
        else:
            # Solo hoja Renglones
            extra = fr.datos_extra or {}
            row = {
                "item": _safe(fr.orden) or str(idx + 1),
                "obj_gasto": _safe(extra.get("obj_gasto", "")) or _safe(extra.get("objeto_gasto", "")) or _safe(extra.get("obj_gas", "")),
                "codigo_item": _safe(fr.codigo_item),
                "descripcion": _safe(fr.descripcion),
                "cantidad": _safe(fr.cantidad),
                "unidad": _safe(fr.unidad),
                "precio_unitario": _safe(extra.get("precio_unitario", "")),
                "importe_total": _safe(extra.get("importe_total", "")),
                "lugar_entrega": _safe(extra.get("lugar_entrega", "")),
                "plazo_entrega": _safe(extra.get("plazo_entrega", "")),
                "periodicidad": _safe(extra.get("periodicidad", "")),
                "marca": _safe(extra.get("marca", "")),
                "destino_efector": _safe(fr.destino_efector),
                "observaciones": _safe(fr.obs_tecnicas),
                "fuente": _safe(extra.get("fuente", "Renglones")),
                "estado": _safe(fr.estado),
            }

        if row["descripcion"] or row["item"]:
            renglones_out.append(row)

    return renglones_out


# ---------------------------------------------------------------------------
# Función principal: calcular_estado_fusion
# ---------------------------------------------------------------------------

def _clean_export_value(value) -> str:
    text = _safe(value)
    return "" if is_fusion_missing(text) else text


def _format_export_date(value) -> str:
    text = _clean_export_value(value)
    if not text:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y %H:%M") if (value.hour or value.minute) else value.strftime("%d/%m/%Y")
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")

    try:
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    except Exception:
        parsed = None
    if parsed is not None and not pd.isna(parsed):
        has_time = bool(getattr(parsed, "hour", 0) or getattr(parsed, "minute", 0))
        return parsed.strftime("%d/%m/%Y %H:%M") if has_time else parsed.strftime("%d/%m/%Y")
    return text


def _format_export_bool(value) -> str:
    text = _clean_export_value(value)
    if not text:
        return ""
    norm = _norm(text)
    token = re.split(r"[\s,;:/()]+", norm)[0] if norm else ""
    if token in {"si", "s", "yes", "true", "1"}:
        return "Si"
    if token in {"no", "false", "0"}:
        return "No"
    return text


def _format_export_value(column: str, value) -> str:
    if column in {
        "Fecha Acto Apertura",
        "Fecha Inicio Consulta",
        "Fecha Final Consulta",
        "Fecha limite para presentación de muestras",
    }:
        return _format_export_date(value)
    if column in {
        "Acepta Redeterminación",
        "Acepta Prórroga",
        "Ampliación",
        "Alternativa",
        "¿Lleva Muestras?",
    }:
        return _format_export_bool(value)
    return _clean_export_value(value)


def _lookup_process_value(caso, proc: dict, *aliases) -> str:
    return _pick(
        _from_proceso(proc, *aliases),
        _from_trazabilidad(caso, *aliases),
        _from_analitica(caso, *aliases),
        _from_hallazgos(caso, *aliases),
    )


def _garantia_value(caso, *keywords) -> str:
    keyword_norms = [_norm(k) for k in keywords]
    for garantia in (caso.garantias or []):
        tipo_norm = _norm(garantia.tipo)
        if not any(k and k in tipo_norm for k in keyword_norms):
            continue
        pieces = []
        for value in (garantia.requerida, garantia.porcentaje, garantia.base_calculo, garantia.plazo, garantia.formas_admitidas):
            clean = _clean_export_value(value)
            if clean:
                pieces.append(clean)
        return " - ".join(pieces) if pieces else _clean_export_value(garantia.tipo)
    return ""


def _lleva_muestras_value(caso, proc: dict) -> str:
    raw = _lookup_process_value(caso, proc, "lleva_muestras", "muestras", "requiere_muestras")
    if raw:
        valor, _ = _normalize_bool_sino(raw)
        return valor
    for requisito in (caso.requisitos or []):
        text = " ".join(_safe(v) for v in (requisito.categoria, requisito.descripcion, requisito.obligatorio))
        if "muestra" in _norm(text):
            return "Si"
    return ""


def _fecha_limite_muestras_value(caso, proc: dict) -> str:
    return _pick(
        _from_cronograma(caso, "muestra", "presentacion de muestras", "presentación de muestras"),
        _lookup_process_value(caso, proc, "fecha_limite_muestras", "fecha_limite_presentacion_muestras", "fecha limite para presentacion de muestras"),
    )


# ---------------------------------------------------------------------------
# Helpers para garantías: porcentaje textual vs. Si/No
# ---------------------------------------------------------------------------

def _format_garantia_porcentaje(valor_raw) -> str:
    """Devuelve texto de porcentaje ('3%', '2,5%') o cadena vacía.
    Nunca devuelve 'No', 'No aplica', marcadores de faltante ni estados internos."""
    text = _clean_export_value(valor_raw)
    if not text:
        return ""
    # Cualquier valor considerado faltante/estado interno → celda vacía
    if is_fusion_missing(text):
        return ""
    n = _norm(text)
    if n in {"no", "false", "0", "no aplica", "no requerida", "no requerido",
             "no corresponde", "no exigida", "no exige", "sin garantia", "sin garantía"}:
        return ""
    if "%" in text:
        return text.strip()
    try:
        num = float(text.replace(",", "."))
        if num == 0:
            return ""
        num_pct = num * 100 if (0 < num < 1) else num
        if num_pct == int(num_pct):
            return f"{int(num_pct)}%"
        return f"{num_pct:.1f}%".replace(".", ",")
    except (ValueError, TypeError):
        pass
    # Si no tiene formato de porcentaje conocido → celda vacía (no exportar texto ambiguo)
    return ""


def _garantia_porcentaje_value(caso, *keywords) -> str:
    """Para garantías de tipo porcentaje: retorna el % como texto o vacío."""
    keyword_norms = [_norm(k) for k in keywords]
    for garantia in (caso.garantias or []):
        tipo_norm = _norm(garantia.tipo)
        if not any(k and k in tipo_norm for k in keyword_norms):
            continue
        requerida = _norm(_clean_export_value(garantia.requerida))
        if requerida in {"no", "false", "0", "no aplica", "no requerida", "no requerido"}:
            return ""
        return _format_garantia_porcentaje(_clean_export_value(garantia.porcentaje))
    return ""


def _garantia_sino_value(caso, *keywords) -> str:
    """Para Garantía de Incorporar Contragarantía: retorna 'Si' o 'No'."""
    keyword_norms = [_norm(k) for k in keywords]
    for garantia in (caso.garantias or []):
        tipo_norm = _norm(garantia.tipo)
        if not any(k and k in tipo_norm for k in keyword_norms):
            continue
        requerida_raw = _clean_export_value(garantia.requerida)
        n = _norm(requerida_raw)
        token = re.split(r"[\s,;:/()]+", n)[0] if n else ""
        if token in {"si", "s", "yes", "true", "1", "sí"}:
            return "Si"
        if token in {"no", "false", "0"}:
            return "No"
        pct = _clean_export_value(garantia.porcentaje)
        if pct and not is_fusion_missing(pct):
            return "Si"
        return ""
    return ""


def _parse_date_native(value):
    """Convierte un valor a datetime nativo de Python para escritura real en Excel."""
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    text = _clean_export_value(value)
    if not text:
        return None
    try:
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
        if parsed is not None and not pd.isna(parsed):
            return parsed.to_pydatetime()
    except Exception:
        pass
    return None


def _resolver_campo_complementario(caso, field_key: str, proc: dict) -> dict:
    """Resuelve campos complementarios para mostrar en la ficha (UI)."""
    valor = ""

    if field_key == "ampliacion":
        raw = _lookup_process_value(caso, proc, "ampliacion", "ampliación", "acepta_ampliacion")
        if raw:
            valor, _ = _normalize_bool_sino(raw)

    elif field_key == "alternativa":
        raw = _lookup_process_value(caso, proc, "alternativa", "oferta_alternativa", "acepta_alternativa")
        if raw:
            valor, _ = _normalize_bool_sino(raw)

    elif field_key == "lugar_entrega":
        valor = _lookup_process_value(
            caso, proc, "lugar_entrega", "lugar y cond de entrega",
            "condiciones_entrega", "lugar_condicion_entrega")

    elif field_key == "exigencia_desestimacion":
        valor = _lookup_process_value(
            caso, proc, "exigencia_desestimacion", "causal_desestimacion",
            "causal de desestimacion", "exigencia con causal")

    elif field_key == "lleva_muestras":
        raw = _lleva_muestras_value(caso, proc)
        if raw:
            valor, _ = _normalize_bool_sino(raw)

    elif field_key == "fecha_limite_muestras":
        valor = _fecha_limite_muestras_value(caso, proc)

    elif field_key == "garantia_mantenimiento_oferta":
        valor = _garantia_porcentaje_value(caso, "mantenimiento", "oferta")

    elif field_key == "garantia_anticipo_financiero":
        valor = _garantia_porcentaje_value(caso, "anticipo")

    elif field_key == "garantia_cumplimiento_contrato":
        valor = _garantia_porcentaje_value(caso, "cumplimiento", "contrato")

    elif field_key == "garantia_impugnacion_pliego":
        valor = _garantia_porcentaje_value(caso, "impugnacion_pliego", "pliego")

    elif field_key == "garantia_impugnacion_preadjudicacion":
        valor = _garantia_porcentaje_value(
            caso, "preadjudicacion", "preadjudicación", "impugnacion_preadjudicacion")

    elif field_key == "garantia_contragarantia":
        valor = _garantia_sino_value(caso, "contragarantia", "contragarantía", "contra garantia")
        if not valor:
            valor = "No"

    completo = bool(valor and not is_fusion_missing(valor))
    return {
        "valor": valor,
        "valor_normalizado": valor,
        "fuente": "Proceso",
        "completo": completo,
        "requiere_validacion": False,
        "estado": "completo" if completo else "faltante",
        "categoria": "pliego_complementario",
    }


def _build_fusion_process_export_row(caso, fusion_ctx: dict) -> dict:
    proc = _get_proceso(caso)
    campos_fusion = fusion_ctx.get("campos_fusion", {})
    row = {}

    _DATE_COLUMNS = {
        "Fecha Acto Apertura",
        "Fecha Inicio Consulta",
        "Fecha Final Consulta",
        "Fecha limite para presentación de muestras",
    }
    _GARANTIA_PCT_COLUMNS = {
        "Garantía de Mantenimiento de Oferta",
        "Garantía de Anticipo Financiero",
        "Garantía de Cumplimiento de Contrato",
        "Garantía de Impugnación al Pliego",
        "Garantía de Impugnación a la Preadjudicación",
    }

    for column in FUSION_EXPORT_PROCESO_COLUMNS:

        # ── ID Proceso: valor manual de la solicitud, nunca ID interno ─────
        if column == "ID Proceso":
            row[column] = _clean_export_value(get_manual_id_proceso(caso)) or None
            continue

        field_key = _FUSION_PROCESS_FIELD_BY_COLUMN.get(column)
        value = ""

        if field_key:
            field_data = campos_fusion.get(field_key, {})
            value = field_data.get("valor_normalizado") or field_data.get("valor") or ""

        # ── Columnas sin field_key: resolución directa ──────────────────────
        elif column == "Ampliación":
            value = _lookup_process_value(caso, proc, "ampliacion", "ampliación", "acepta_ampliacion")
        elif column == "Alternativa":
            value = _lookup_process_value(caso, proc, "alternativa", "oferta_alternativa", "acepta_alternativa")
        elif column == "Lugar y Cond. De Entrega":
            value = _lookup_process_value(
                caso, proc, "lugar_entrega", "lugar y cond de entrega",
                "condiciones_entrega", "plazo_entrega", "periodicidad")

        # ── Garantías porcentuales: texto "3%" o celda vacía ───────────────
        elif column == "Garantía de Mantenimiento de Oferta":
            value = _garantia_porcentaje_value(caso, "mantenimiento", "oferta")
        elif column == "Garantía de Anticipo Financiero":
            value = _garantia_porcentaje_value(caso, "anticipo")
        elif column == "Garantía de Cumplimiento de Contrato":
            value = _garantia_porcentaje_value(caso, "cumplimiento", "contrato")
        elif column == "Garantía de Impugnación al Pliego":
            # "pliego" no aparece en 'impugnacion_preadjudicacion' → keyword específica
            value = _garantia_porcentaje_value(caso, "impugnacion_pliego", "pliego")
        elif column == "Garantía de Impugnación a la Preadjudicación":
            value = _garantia_porcentaje_value(
                caso, "preadjudicacion", "preadjudicación", "impugnacion_preadjudicacion")

        # ── Contragarantía: Si/No ───────────────────────────────────────────
        elif column == "Garantía de Incorporar Contragarantía":
            value = _garantia_sino_value(
                caso, "contragarantia", "contragarantía", "contra garantia")
            if not value:
                value = "No"

        elif column == "Exigencia con Causal de Desestimación":
            value = _lookup_process_value(
                caso, proc, "exigencia_desestimacion", "causal_desestimacion",
                "causal de desestimacion")
        elif column == "¿Lleva Muestras?":
            value = _lleva_muestras_value(caso, proc)
        elif column == "Fecha limite para presentación de muestras":
            value = _fecha_limite_muestras_value(caso, proc)

        # ── Monto: número nativo (sin símbolo de moneda) ───────────────────
        if column == "Monto" and value:
            try:
                clean_num = re.sub(r"[^\d.,]", "", str(value))
                if "," in clean_num and "." not in clean_num:
                    clean_num = clean_num.replace(",", ".")
                elif "," in clean_num and "." in clean_num:
                    clean_num = clean_num.replace(",", "")
                row[column] = float(clean_num)
                continue
            except (ValueError, TypeError):
                pass

        # ── Fechas: datetime nativo (serial Excel real) ─────────────────────
        if column in _DATE_COLUMNS:
            row[column] = _parse_date_native(value)
            continue

        # ── Garantías porcentuales: guardar como texto puro (no como número)
        if column in _GARANTIA_PCT_COLUMNS:
            row[column] = _clean_export_value(value) or None
            continue

        # ── Resto: formatear con reglas estándar ───────────────────────────
        row[column] = _format_export_value(column, value)

    return row


def _build_fusion_renglones_export_rows(caso, fusion_ctx: dict) -> list[dict]:
    rows = []
    for row in fusion_ctx.get("renglones_fusion", []) or []:
        rows.append({
            "Item": _clean_export_value(row.get("item")),
            "Obj. Gas.": None,
            "Cod. Item": None,
            "Descripción": _clean_export_value(row.get("descripcion")),
            "Cant": _clean_export_value(row.get("cantidad")),
        })
    return rows


def _validate_fusion_template_workbook(workbook) -> None:
    expected_sheets = ["Datos del proceso", "Detalle"]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"Plantilla Fusion inválida: hojas {workbook.sheetnames}, esperado {expected_sheets}")

    ws_proc = workbook["Datos del proceso"]
    headers_proc = [ws_proc.cell(1, col).value for col in range(1, len(FUSION_PROCESS_FIELDS) + 1)]
    if headers_proc != FUSION_PROCESS_FIELDS:
        raise ValueError("Plantilla Fusion inválida: encabezados de 'Datos del proceso' no coinciden con A:AF")

    ws_det = workbook["Detalle"]
    headers_det = [ws_det.cell(1, col).value for col in range(1, len(FUSION_EXPORT_RENGLON_COLUMNS) + 1)]
    if headers_det != FUSION_EXPORT_RENGLON_COLUMNS:
        raise ValueError("Plantilla Fusion inválida: encabezados de 'Detalle' no coinciden con A:E")


def _clear_template_values(ws, min_row: int, max_col: int) -> None:
    for row in range(min_row, max(ws.max_row, min_row) + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def export_fusion_excel_bytes(
    caso,
    audit_history: list | None = None,
    fusion_ctx_override: dict | None = None,
) -> bytes:
    """Genera el Excel final para carga en Fusion/ERP.
    Estructura idéntica al Documento_Importación 2.xlsx (plantilla maestra).
    Si se pasa fusion_ctx_override, usa ese contexto (con overrides ya aplicados).
    Si se pasa audit_history (lista de dicts), agrega hoja 'Auditoria_Ediciones'.
    """
    fusion_ctx = fusion_ctx_override if fusion_ctx_override is not None else calcular_estado_fusion(caso)
    proceso_row = _build_fusion_process_export_row(caso, fusion_ctx)
    renglones_rows = _build_fusion_renglones_export_rows(caso, fusion_ctx)

    if not FUSION_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"No se encontró la plantilla Fusion: {FUSION_TEMPLATE_PATH}")

    workbook = openpyxl.load_workbook(FUSION_TEMPLATE_PATH)
    _validate_fusion_template_workbook(workbook)

    ws_proc = workbook["Datos del proceso"]
    ws_det = workbook["Detalle"]

    _clear_template_values(ws_proc, min_row=2, max_col=len(FUSION_EXPORT_PROCESO_COLUMNS))
    _clear_template_values(ws_det, min_row=2, max_col=len(FUSION_EXPORT_RENGLON_COLUMNS))

    for col_idx, column in enumerate(FUSION_EXPORT_PROCESO_COLUMNS, start=1):
        cell = ws_proc.cell(row=2, column=col_idx)
        cell.value = proceso_row.get(column)
        if column in {
            "Fecha Inicio Consulta",
            "Fecha Final Consulta",
            "Fecha Acto Apertura",
            "Fecha limite para presentación de muestras",
        }:
            cell.number_format = "dd/mm/yyyy\\ hh:mm:ss"
        elif column == "Monto":
            cell.number_format = "#,##0.00"
        elif column in {
            "Garantía de Mantenimiento de Oferta",
            "Garantía de Anticipo Financiero",
            "Garantía de Cumplimiento de Contrato",
            "Garantía de Impugnación al Pliego",
            "Garantía de Impugnación a la Preadjudicación",
        }:
            cell.number_format = "@"

    for row_idx, row_data in enumerate(renglones_rows, start=2):
        if row_idx > 2:
            _copy_row_style(ws_det, source_row=2, target_row=row_idx, max_col=len(FUSION_EXPORT_RENGLON_COLUMNS))
        for col_idx, column in enumerate(FUSION_EXPORT_RENGLON_COLUMNS, start=1):
            ws_det.cell(row=row_idx, column=col_idx).value = row_data.get(column)

    # Hoja de auditoría de ediciones manuales
    if audit_history:
        _append_audit_sheet(workbook, audit_history)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


_AUDIT_HEADERS = [
    "Fecha y hora", "Usuario", "Rol", "Sección", "Campo", "Clave campo",
    "Valor anterior", "Valor nuevo", "Estado anterior", "Estado nuevo", "Motivo",
]


def _append_audit_sheet(workbook: openpyxl.Workbook, audit_history: list) -> None:
    """Agrega la hoja Auditoria_Ediciones al workbook."""
    ws = workbook.create_sheet("Auditoria_Ediciones")

    # Estilos de encabezado
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(_AUDIT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    alt_fill = PatternFill("solid", fgColor="EBF3FF")
    for row_idx, h in enumerate(audit_history, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            h.get("edited_at", ""),
            h.get("edited_by_name", ""),
            h.get("edited_by_role", ""),
            h.get("section_key", ""),
            h.get("field_label", ""),
            h.get("field_key", ""),
            h.get("old_value", ""),
            h.get("new_value", ""),
            h.get("old_status", ""),
            h.get("new_status", ""),
            h.get("reason", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    # Ancho de columnas
    col_widths = [20, 22, 12, 20, 24, 24, 30, 30, 18, 18, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 22


def calcular_estado_fusion(caso) -> dict:
    """
    Calcula el estado Fusion del caso de forma independiente a LicIA.
    Retorna:
    - estado_fusion: 'listo_para_fusion' | 'no_listo_para_fusion' | 'requiere_validacion_manual'
    - estado_visualizacion: 'listo_para_visualizacion' | 'requiere_revision' | 'en_reproceso' | 'incompleto'
    - estado_licia: valor informativo leído del Excel
    - faltantes_criticos: lista de campos obligatorios faltantes
    - campos_fusion: dict con todos los campos resueltos y sus estados
    - renglones_fusion: lista de renglones procesados
    - contradiccion_licia: bool — LicIA dice listo pero SIEM detecta faltantes
    - mensaje_contradiccion: str
    - resumen: dict con conteos
    """
    fc = _get_fusion_cabecera(caso)
    proc = _get_proceso(caso)
    ctrl = _get_control_carga(caso)

    # Estado informativo de LicIA
    licia_listo = _safe(
        fc.get("listo_para_fusion") or ctrl.get("listo_para_fusion")
    )
    licia_listo_visualizacion = _safe(
        fc.get("listo_para_visualizacion") or ctrl.get("listo_para_visualizacion")
    )
    licia_requiere_revision = _safe(
        fc.get("requiere_revision_analista") or ctrl.get("requiere_revision_analista")
    )
    estado_licia = {
        "listo_para_fusion": licia_listo,
        "listo_para_visualizacion": licia_listo_visualizacion,
        "requiere_revision_analista": licia_requiere_revision,
        "nivel_riesgo": _safe(ctrl.get("nivel_riesgo_integracion", "")),
        "motivo_estado": _safe(ctrl.get("motivo_estado", "")),
        "accion_recomendada": _safe(ctrl.get("accion_recomendada", "")),
        "faltantes_documentales": _safe(ctrl.get("faltantes_documentales", "")),
        "contradicciones_detectadas": _safe(ctrl.get("contradicciones_detectadas", "")),
    }

    # Resolver cada campo del modelo Fusion/ERP (32 columnas A:AF)
    campos_fusion = {}
    faltantes_criticos = []
    campos_completos = 0
    requiere_validacion_manual = False

    for campo_meta in FUSION_CAMPOS_OBLIGATORIOS:
        key = campo_meta["key"]
        if key == "id_proceso":
            manual_id_proceso = get_manual_id_proceso(caso)
            resultado = {
                "valor": manual_id_proceso,
                "valor_normalizado": manual_id_proceso,
                "fuente": "Solicitud",
                "completo": bool(manual_id_proceso),
                "requiere_validacion": False,
            }
        elif key in _FUSION_CORE_FIELD_KEYS:
            resultado = _resolver_campo_fusion(caso, key, fc, proc)
        else:
            resultado = _resolver_campo_complementario(caso, key, proc)
        resultado["key"] = key
        resultado["label"] = campo_meta["label"]
        resultado["categoria"] = campo_meta["categoria"]

        if resultado["completo"]:
            campos_completos += 1
            if resultado["requiere_validacion"]:
                requiere_validacion_manual = True
                resultado["estado"] = "requiere_validacion_manual"
            else:
                resultado["estado"] = "completo"
        else:
            resultado["estado"] = "faltante_critico_fusion"
            faltantes_criticos.append({
                "key": key,
                "label": campo_meta["label"],
                "categoria": campo_meta["categoria"],
                "hoja_origen": resultado.get("fuente", "—"),
            })

        campos_fusion[key] = resultado

    # Renglones
    renglones_fusion = _resolver_renglones_fusion(caso)
    tiene_renglones = len(renglones_fusion) > 0

    # Calcular estado_fusion SIEM (independiente de LicIA)
    total_obligatorios = len(FUSION_CAMPOS_OBLIGATORIOS)
    if not faltantes_criticos and tiene_renglones and not requiere_validacion_manual:
        estado_fusion = "listo_para_fusion"
    elif requiere_validacion_manual or (faltantes_criticos and campos_completos >= total_obligatorios - 3):
        estado_fusion = "requiere_validacion_manual"
    else:
        estado_fusion = "no_listo_para_fusion"

    # Estado visualización
    if caso.datos_proceso and (caso.renglones or caso.fusion_renglones):
        if faltantes_criticos or requiere_validacion_manual:
            estado_visualizacion = "requiere_revision"
        else:
            estado_visualizacion = "listo_para_visualizacion"
    elif caso.datos_proceso:
        estado_visualizacion = "requiere_revision"
    else:
        estado_visualizacion = "incompleto"

    # Detectar contradicción con LicIA
    licia_dice_listo = _norm(licia_listo) in {"si", "sí", "yes", "true", "1"}
    contradiccion_licia = licia_dice_listo and bool(faltantes_criticos)
    mensaje_contradiccion = (
        "El Excel de LicIA indica listo para Fusion, pero SIEM detectó faltantes obligatorios según la matriz Fusion."
        if contradiccion_licia else ""
    )

    # Clasificar faltantes/dudas para separar documentales de campos Fusion
    faltantes_documentales = []
    faltantes_campos_fusion = []
    dudas_operativas = []

    for f in (caso.faltantes or []):
        campo_norm = _norm(f.campo_objetivo)
        es_campo_fusion = any(
            campo_norm == _norm(c["key"]) or campo_norm in _norm(c["key"]) or _norm(c["key"]) in campo_norm
            for c in FUSION_CAMPOS_OBLIGATORIOS
        )
        criticidad_norm = _norm(f.criticidad or "")
        if "documental" in (f.motivo or "").lower() or "documento" in campo_norm:
            faltantes_documentales.append(f)
        elif es_campo_fusion and criticidad_norm in {"alta", "critica", "crítica", "alto", "high"}:
            faltantes_campos_fusion.append(f)
        else:
            dudas_operativas.append(f)

    return {
        "estado_fusion": estado_fusion,
        "estado_visualizacion": estado_visualizacion,
        "estado_licia": estado_licia,
        "estado_licia_texto": _build_licia_texto(estado_licia),
        "faltantes_criticos": faltantes_criticos,
        "campos_fusion": campos_fusion,
        "campos_completos": campos_completos,
        "campos_total_obligatorios": total_obligatorios,
        "renglones_fusion": renglones_fusion,
        "tiene_renglones": tiene_renglones,
        "cantidad_renglones": len(renglones_fusion),
        "requiere_validacion_manual": requiere_validacion_manual,
        "contradiccion_licia": contradiccion_licia,
        "mensaje_contradiccion": mensaje_contradiccion,
        "faltantes_documentales": faltantes_documentales,
        "faltantes_campos_fusion": faltantes_campos_fusion,
        "dudas_operativas": dudas_operativas,
        "resumen": {
            # aliases usados en templates
            "campos_completos": campos_completos,
            "campos_totales": total_obligatorios,
            "renglones": len(renglones_fusion),
            # claves completas
            "campos_obligatorios_completos": campos_completos,
            "campos_obligatorios_faltantes": len(faltantes_criticos),
            "campos_requieren_validacion": sum(1 for c in campos_fusion.values() if c.get("estado") == "requiere_validacion_manual"),
            "cantidad_renglones": len(renglones_fusion),
            "faltantes_documentales": len(faltantes_documentales),
            "requiere_revision_analista": bool(licia_requiere_revision and _norm(licia_requiere_revision) in {"si", "sí", "yes", "true", "1"}),
            "contradicciones_detectadas": _safe(ctrl.get("contradicciones_detectadas", "0")),
        },
    }
