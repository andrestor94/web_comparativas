"""Resolucion auditable de cuentas CRM candidatas para oportunidades.

La fuente de identidad es el CUIT del run de Dimensionamiento. El maestro
``clientes.csv`` no contiene CUIT: sus codigos se vinculan unicamente mediante el
nombre legal canonico observado para ese CUIT. ``Operadores.xlsx`` enriquece el
codigo, pero nunca decide el usuario asignado de SuiteCRM.
"""
from __future__ import annotations

import csv
import math
import re
import threading
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .identity import canon
from .models import DimensionamientoRecord, OportunidadSummary

_NULL_MARKERS = {"", "SIN DATO", "SIN_DATO", "S/D", "NULL", "NONE", "NAN", "0", "-"}
_DATA_DIR = Path(__file__).resolve().parents[1] / "data"
CLIENTES_PATH = _DATA_DIR / "forecast_data" / "clientes.csv"
OPERADORES_PATH = _DATA_DIR / "archivos dimensionamiento" / "Operadores.xlsx"
MAX_CANDIDATES = 25


def normalize_identifier(value: Any) -> str | None:
    """Normaliza cuentas/codigos como texto sin perder ceros iniciales."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            return None
        text = format(value, ".15g")
    else:
        text = str(value).strip()
    if not text or text.upper() in _NULL_MARKERS:
        return None
    if re.fullmatch(r"[+-]?\d+\.0+", text):
        text = text.split(".", 1)[0]
    elif re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+", text):
        try:
            decimal = Decimal(text)
        except InvalidOperation:
            return None
        if decimal != decimal.to_integral_value():
            return None
        text = format(decimal.quantize(Decimal(1)), "f")
    text = text.strip()
    return text if text and text.upper() not in _NULL_MARKERS else None


def normalize_cuit(value: Any) -> str | None:
    """Normaliza CUIT comparables sin completar ni inventar digitos."""
    text = normalize_identifier(value)
    if not text:
        return None
    digits = "".join(re.findall(r"\d", text))
    return digits if 10 <= len(digits) <= 11 else None


@dataclass(frozen=True)
class MasterIndex:
    clients_by_name: dict[str, tuple[dict[str, str], ...]]
    operators_by_code: dict[str, dict[str, str | None]]
    signature: tuple[Any, ...]


_master_lock = threading.Lock()
_master_cache: MasterIndex | None = None


def _file_signature(*paths: Path) -> tuple[Any, ...]:
    signature: list[Any] = []
    for path in paths:
        stat = path.stat()
        signature.extend((str(path), stat.st_mtime_ns, stat.st_size))
    return tuple(signature)


def _load_master_index(clientes_path: Path, operadores_path: Path) -> MasterIndex:
    clients: dict[str, list[dict[str, str]]] = {}
    with clientes_path.open(encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        required = {"codigo", "nombre", "fantasia", "grupo", "cliente_grupo", "nombre_grupo", "tipocli"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise ValueError("clientes.csv no tiene el esquema comercial esperado.")
        for raw in reader:
            code = normalize_identifier(raw.get("codigo"))
            name_key = canon(raw.get("nombre"))
            if not code or not name_key:
                continue
            row = {key: (raw.get(key) or "").strip() for key in required}
            row["codigo"] = code
            clients.setdefault(name_key, []).append(row)

    workbook = load_workbook(operadores_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required_xlsx = {"Codigo", "Fantasia", "Vendedor", "Nombre"}
    if not required_xlsx.issubset(set(headers)):
        workbook.close()
        raise ValueError("Operadores.xlsx no tiene el esquema esperado.")
    operators: dict[str, dict[str, str | None]] = {}
    for values in rows:
        raw = dict(zip(headers, values))
        code = normalize_identifier(raw.get("Codigo"))
        if not code:
            continue
        operators[code] = {
            "codigo_cuenta": code,
            "fantasia": str(raw.get("Fantasia") or "").strip() or None,
            "vendedor_codigo": normalize_identifier(raw.get("Vendedor")),
            "nombre": str(raw.get("Nombre") or "").strip() or None,
        }
    workbook.close()
    return MasterIndex(
        clients_by_name={key: tuple(value) for key, value in clients.items()},
        operators_by_code=operators,
        signature=_file_signature(clientes_path, operadores_path),
    )


def get_master_index(
    clientes_path: Path = CLIENTES_PATH,
    operadores_path: Path = OPERADORES_PATH,
) -> MasterIndex:
    """Indice inmutable; solo publica una recarga completa y estable."""
    global _master_cache
    with _master_lock:
        signature_before = _file_signature(clientes_path, operadores_path)
        if _master_cache is None or _master_cache.signature != signature_before:
            loaded = _load_master_index(clientes_path, operadores_path)
            if loaded.signature != signature_before:
                raise ValueError(
                    "Los maestros cambiaron durante la lectura; reintentá cuando finalice la actualización."
                )
            _master_cache = loaded
        return _master_cache


def reset_master_cache() -> None:
    global _master_cache
    with _master_lock:
        _master_cache = None


def _add_candidate(
    ordered: list[dict[str, Any]],
    by_code: dict[str, dict[str, Any]],
    code: Any,
    *,
    source: str,
    master_row: dict[str, str] | None = None,
    operator: dict[str, str | None] | None = None,
    confidence: str = "directa",
) -> None:
    normalized = normalize_identifier(code)
    if not normalized:
        return
    candidate = by_code.get(normalized)
    if candidate is None:
        candidate = {
            "cuenta": normalized,
            "codigo_relacionado": normalized,
            "cliente_maestro": (master_row or {}).get("nombre") or None,
            "fantasia": (master_row or {}).get("fantasia") or None,
            "operador_codigo": (operator or {}).get("vendedor_codigo"),
            "operador_nombre": (operator or {}).get("nombre"),
            "fuentes": [],
            "relacion_confianza": confidence,
            "existe_en_crm": None,
            "crm_account_id": None,
            "crm_nombre": None,
        }
        by_code[normalized] = candidate
        ordered.append(candidate)
    if confidence == "confirmada_por_cuit_dataset":
        candidate["relacion_confianza"] = confidence
    if source not in candidate["fuentes"]:
        candidate["fuentes"].append(source)
    if master_row and not candidate.get("cliente_maestro"):
        candidate["cliente_maestro"] = master_row.get("nombre") or None
        candidate["fantasia"] = master_row.get("fantasia") or None
    if operator and not candidate.get("operador_nombre"):
        candidate["operador_codigo"] = operator.get("vendedor_codigo")
        candidate["operador_nombre"] = operator.get("nombre")


_GENERIC_NAME_TOKENS = {
    "HOSPITAL", "CLINICA", "SANATORIO", "INSTITUTO", "CENTRO", "FUNDACION",
    "SOCIEDAD", "ASOCIACION", "MEDICA", "MEDICO", "SALUD", "SERVICIOS",
    "PRIVADO", "PRIVADA", "SA", "SRL", "SAS", "S", "A",
}


@dataclass(frozen=True)
class RunIdentityIndex:
    names_by_cuit: dict[str, tuple[str, ...]]
    cuits_by_name: dict[str, tuple[str, ...]]
    raw_names_by_key: dict[str, tuple[str, ...]]
    accounts_by_cuit: dict[str, tuple[str, ...]]


_identity_lock = threading.Lock()
_identity_cache: dict[tuple[str, int], RunIdentityIndex] = {}


def _build_run_identity_index(db: Session, run_id: int) -> RunIdentityIndex:
    names_by_cuit: dict[str, set[str]] = {}
    cuits_by_name: dict[str, set[str]] = {}
    raw_names_by_key: dict[str, set[str]] = {}
    accounts_by_cuit: dict[str, set[str]] = {}
    rows = db.execute(
        select(
            DimensionamientoRecord.cuit,
            DimensionamientoRecord.cliente_nombre_homologado,
            DimensionamientoRecord.cuenta_interna,
        )
        .where(DimensionamientoRecord.import_run_id == run_id)
        .distinct()
    ).all()
    for raw_cuit, raw_name, raw_account in rows:
        cuit = normalize_cuit(raw_cuit)
        key = canon(raw_name)
        if not cuit or not key or key == "SIN DATO":
            continue
        names_by_cuit.setdefault(cuit, set()).add(key)
        cuits_by_name.setdefault(key, set()).add(cuit)
        raw_names_by_key.setdefault(key, set()).add(str(raw_name).strip())
        account = normalize_identifier(raw_account)
        if account:
            accounts_by_cuit.setdefault(cuit, set()).add(account)
    return RunIdentityIndex(
        names_by_cuit={key: tuple(sorted(value)) for key, value in names_by_cuit.items()},
        cuits_by_name={key: tuple(sorted(value)) for key, value in cuits_by_name.items()},
        raw_names_by_key={key: tuple(sorted(value)) for key, value in raw_names_by_key.items()},
        accounts_by_cuit={key: tuple(sorted(value)) for key, value in accounts_by_cuit.items()},
    )


def get_run_identity_index(db: Session, run_id: int) -> RunIdentityIndex:
    bind = db.get_bind()
    key = (str(getattr(bind, "url", bind)), int(run_id))
    with _identity_lock:
        cached = _identity_cache.get(key)
        if cached is None:
            cached = _build_run_identity_index(db, run_id)
            _identity_cache[key] = cached
        return cached


def reset_identity_cache() -> None:
    with _identity_lock:
        _identity_cache.clear()


def _is_generic_name_key(key: str) -> bool:
    """Bloquea claves formadas solo por términos institucionales o siglas cortas."""
    tokens = [token for token in key.split() if token]
    informative = [
        token for token in tokens
        if token not in _GENERIC_NAME_TOKENS and len(token) > 2
    ]
    return not informative


def relationship_candidates(
    db: Session,
    run_id: int,
    opportunity: OportunidadSummary,
    *,
    master: MasterIndex | None = None,
    identity: RunIdentityIndex | None = None,
) -> dict[str, Any]:
    """Construye alternativas solo con igualdad exacta de razón social normalizada.

    El puente es ``dimensionamiento_records.cliente_nombre_homologado`` -> ``canon``
    -> ``clientes.nombre`` -> ``clientes.codigo``. No usa ``contains``, prefijos,
    similitud, fuzzy matching, grupos comerciales ni elimina sufijos societarios.
    """
    master = master or get_master_index()
    identity = identity or get_run_identity_index(db, run_id)
    cuit = normalize_cuit(getattr(opportunity, "cuit", None))
    original = normalize_identifier(getattr(opportunity, "cuenta_interna", None))
    warnings: list[str] = []
    ambiguity_reasons: list[str] = []
    ordered: list[dict[str, Any]] = []
    by_code: dict[str, dict[str, Any]] = {}
    _add_candidate(
        ordered, by_code, original, source="cuenta_original_dataset",
        operator=master.operators_by_code.get(original or ""), confidence="cuenta_original",
    )

    names = list(identity.names_by_cuit.get(cuit or "", ()))
    if not cuit:
        warnings.append("El CUIT es nulo o inválido; no se construyeron alternativas por nombre.")
    elif not names:
        warnings.append("El CUIT no tiene cliente_nombre_homologado utilizable en la corrida.")
    elif len(names) > 1:
        ambiguity_reasons.append("CUIT_CON_MULTIPLES_NOMBRES_HOMOLOGADOS")

    for name in names:
        linked_cuits = identity.cuits_by_name.get(name, ())
        if len(linked_cuits) > 1:
            ambiguity_reasons.append("NOMBRE_CANONICO_COMPARTIDO_POR_MULTIPLES_CUIT")
        if _is_generic_name_key(name):
            ambiguity_reasons.append("CLAVE_NOMINAL_SIN_TERMINOS_DISTINTIVOS")

    # El conteo auditado incluye todo el universo potencial aun cuando una colisión
    # impida convertir esas cuentas en opciones seleccionables. Así la UI puede decir
    # "22 detectadas" sin exponerlas ni consultarlas como si fueran relaciones válidas.
    potential_accounts = {original} if original else set()
    if cuit:
        potential_accounts.update(identity.accounts_by_cuit.get(cuit, ()))
    for name in names:
        potential_accounts.update(row["codigo"] for row in master.clients_by_name.get(name, ()))

    # Una colisión nominal bloquea todo el puente. No se elige la primera cuenta ni se
    # usa frecuencia/orden como desempate.
    if cuit and not ambiguity_reasons:
        for account in identity.accounts_by_cuit.get(cuit, ()):
            _add_candidate(
                ordered, by_code, account, source="dataset_cuenta_mismo_cuit",
                operator=master.operators_by_code.get(account),
                confidence="confirmada_por_cuit_dataset",
            )
        for name in names:
            for row in master.clients_by_name.get(name, ()):
                code = row["codigo"]
                _add_candidate(
                    ordered, by_code, code,
                    source="clientes_nombre_homologado_exacto_no_ambiguo",
                    master_row=row, operator=master.operators_by_code.get(code),
                    confidence="relacionada_por_nombre_exacto_no_ambiguo",
                )

    candidate_count = len(potential_accounts)
    too_many = candidate_count > MAX_CANDIDATES
    if too_many:
        ambiguity_reasons.append("MAS_DE_25_CUENTAS_CANDIDATAS")
        warnings.append(
            f"Se encontraron {candidate_count} cuentas candidatas. No se truncaron ni consultaron; requiere revisión."
        )
    ambiguous = bool(ambiguity_reasons)
    return {
        "cuit": cuit,
        "cuenta_original": original,
        "nombres_homologados_canonicos": names,
        "nombres_homologados_originales": sorted({
            raw for name in names for raw in identity.raw_names_by_key.get(name, ())
        }),
        "cuentas_candidatas": ordered,
        "cantidad_candidatas_total": candidate_count,
        "fuente_relacion": (
            "dataset.cliente_nombre_homologado -> canon exacto -> "
            "clientes.nombre -> clientes.codigo -> Operadores.Codigo"
        ),
        "coincidencia_nombre": "igualdad_exacta",
        "relacion_valida": bool(cuit and names and not ambiguous),
        "relacion_ambigua": ambiguous,
        "motivos_ambiguedad": sorted(set(ambiguity_reasons)),
        "exceso_candidatos": too_many,
        "advertencias": warnings,
    }

class AccountSelectionError(ValueError):
    pass


_CONFIDENCE_LABELS = {
    "CUENTA_ORIGINAL_CONFIRMADA": "Cuenta original confirmada en CRM",
    "ALTERNATIVA_CONFIRMADA_POR_CUIT": "Cuenta alternativa confirmada por CUIT",
    "ALTERNATIVA_RELACIONADA_POR_NOMBRE_EXACTO_NO_AMBIGUO": (
        "Cuenta alternativa relacionada por razón social normalizada; verificar antes de enviar"
    ),
    "RELACION_AMBIGUA": "Relación ambigua; requiere revisión",
    "SIN_RELACION": "Sin relación de cuenta utilizable",
    "ERROR_CONSULTA_CRM": "Error de consulta CRM",
}


def _trace_text(
    original: str | None,
    selected: dict[str, Any],
    cuit: str | None,
    criterion: str,
    confidence: str,
) -> str:
    final = selected["cuenta"]
    if confidence == "CUENTA_ORIGINAL_CONFIRMADA":
        return f"Cuenta CRM {final} validada como cuenta original del dataset."
    if confidence == "ALTERNATIVA_CONFIRMADA_POR_CUIT":
        return (
            f"Cuenta original {original or 'sin dato'}; cuenta CRM utilizada {final}; "
            f"alternativa vinculada al CUIT {cuit or 'sin dato'}. Criterio: {criterion}."
        )
    return (
        f"Cuenta original {original or 'sin dato'}; cuenta CRM utilizada {final}; "
        "alternativa relacionada únicamente por igualdad exacta de razón social normalizada. "
        "El CRM no confirmó identidad fiscal. Verificada manualmente antes del envío. "
        f"Criterio: {criterion}."
    )


def _base_resolution(
    relationship: dict[str, Any],
    candidates: list[dict[str, Any]],
    *,
    crm_mode: str,
    confidence: str,
    warnings: list[str],
) -> dict[str, Any]:
    return {
        **relationship,
        "crm_modo": crm_mode,
        "cuentas_candidatas": candidates,
        "cuentas_encontradas_en_crm": [],
        "cuenta_seleccionada": None,
        "criterio_seleccion": confidence.lower(),
        "estado_confianza": confidence,
        "confianza_label": _CONFIDENCE_LABELS[confidence],
        "confirmacion_fiscal": False,
        "seleccion_origen": "ninguna",
        "requiere_seleccion": False,
        "bloqueado": True,
        "advertencias": warnings,
        "trazabilidad_texto": None,
    }


def select_account_resolution(
    relationship: dict[str, Any],
    crm_results: dict[str, dict[str, Any]],
    *,
    crm_mode: str,
    requested_account: Any = None,
) -> dict[str, Any]:
    """Aplica confianza y selección sin convertir coincidencia nominal en identidad fiscal."""
    candidates = [dict(row) for row in relationship.get("cuentas_candidatas") or []]
    by_code = {row["cuenta"]: row for row in candidates}
    requested = normalize_identifier(requested_account)
    if requested and requested not in by_code:
        raise AccountSelectionError(
            f"La cuenta {requested} no pertenece a las cuentas relacionadas con la oportunidad."
        )
    warnings = list(relationship.get("advertencias") or [])
    original = relationship.get("cuenta_original")
    cuit = relationship.get("cuit")

    if relationship.get("relacion_ambigua") or relationship.get("exceso_candidatos"):
        return _base_resolution(
            relationship, candidates, crm_mode=crm_mode,
            confidence="RELACION_AMBIGUA", warnings=warnings,
        )

    if crm_mode == "simulado":
        selected = by_code.get(original)
        if not selected:
            return _base_resolution(
                relationship, candidates, crm_mode=crm_mode,
                confidence="SIN_RELACION", warnings=warnings,
            )
        confidence = "CUENTA_ORIGINAL_CONFIRMADA"
        criterion = "simulado_cuenta_original"
        trace = _trace_text(original, selected, cuit, criterion, confidence)
        return {
            **relationship, "crm_modo": crm_mode, "cuentas_candidatas": candidates,
            "cuentas_encontradas_en_crm": [], "cantidad_evaluadas_crm": 0,
            "cuenta_seleccionada": selected,
            "criterio_seleccion": criterion, "estado_confianza": confidence,
            "confianza_label": _CONFIDENCE_LABELS[confidence], "confirmacion_fiscal": False,
            "seleccion_origen": "automatica_original", "requiere_seleccion": False,
            "bloqueado": False, "advertencias": warnings, "trazabilidad_texto": trace,
        }

    lookup_errors = False
    fiscal_mismatch = False
    found: list[dict[str, Any]] = []
    for candidate in candidates:
        result = crm_results.get(candidate["cuenta"], {})
        candidate["existe_en_crm"] = result.get("exists")
        candidate["crm_account_id"] = result.get("crm_account_id")
        candidate["crm_nombre"] = result.get("crm_nombre")
        candidate["crm_numero_cuenta"] = result.get("crm_numero_cuenta")
        candidate["crm_cuit"] = normalize_cuit(result.get("crm_cuit"))
        candidate["crm_razon_social"] = result.get("crm_razon_social")
        candidate["crm_documento"] = result.get("crm_documento")
        if result.get("error"):
            candidate["error_consulta"] = result["error"]
            lookup_errors = True
        if result.get("exists") is not True:
            continue
        is_original = candidate["cuenta"] == original
        crm_cuit = candidate.get("crm_cuit")
        if crm_cuit and cuit and crm_cuit != cuit:
            candidate["identidad_fiscal_divergente"] = True
            fiscal_mismatch = True
            warnings.append(
                f"La cuenta {candidate['cuenta']} devolvió CUIT {crm_cuit}, distinto de {cuit}."
            )
            continue
        if is_original:
            confidence = "CUENTA_ORIGINAL_CONFIRMADA"
        elif crm_cuit and cuit and crm_cuit == cuit:
            confidence = "ALTERNATIVA_CONFIRMADA_POR_CUIT"
        else:
            confidence = "ALTERNATIVA_RELACIONADA_POR_NOMBRE_EXACTO_NO_AMBIGUO"
        candidate["estado_confianza"] = confidence
        candidate["confianza_label"] = _CONFIDENCE_LABELS[confidence]
        candidate["confirmacion_fiscal"] = bool(crm_cuit and cuit and crm_cuit == cuit)
        candidate["trazabilidad_seleccion"] = _trace_text(
            original, candidate, cuit, "seleccion_manual_entre_alternativas", confidence,
        )
        found.append(candidate)

    if lookup_errors:
        warnings.append("La consulta al CRM fue parcial; no se aplicó ningún fallback.")
        return _base_resolution(
            relationship, candidates, crm_mode=crm_mode,
            confidence="ERROR_CONSULTA_CRM", warnings=warnings,
        )
    if fiscal_mismatch:
        relationship = {
            **relationship,
            "relacion_ambigua": True,
            "motivos_ambiguedad": sorted(set(
                list(relationship.get("motivos_ambiguedad") or [])
                + ["CUIT_CRM_DIVERGENTE"]
            )),
        }
        return _base_resolution(
            relationship, candidates, crm_mode=crm_mode,
            confidence="RELACION_AMBIGUA", warnings=warnings,
        )

    original_found = next((row for row in found if row["cuenta"] == original), None)
    if original_found:
        if requested and requested != original:
            raise AccountSelectionError("La cuenta original existe en el CRM y debe conservarse.")
        selected = original_found
        criterion = "cuenta_original_existente"
        origin = "automatica_original"
        requires = False
    else:
        alternatives = [row for row in found if row["cuenta"] != original]
        if len(alternatives) == 1:
            selected = alternatives[0]
            if requested and requested != selected["cuenta"]:
                raise AccountSelectionError("La cuenta elegida no es la única alternativa válida en el CRM.")
            criterion = "unica_alternativa_valida"
            origin = "automatica_unica_alternativa"
            requires = False
        elif len(alternatives) > 1:
            selected = by_code.get(requested) if requested else None
            if selected and selected not in alternatives:
                raise AccountSelectionError("La cuenta elegida no es una alternativa válida en el CRM.")
            criterion = "seleccion_manual_entre_alternativas" if selected else "multiples_alternativas"
            origin = "manual" if selected else "ninguna"
            requires = selected is None
        else:
            selected = None
            criterion = "sin_relacion"
            origin = "ninguna"
            requires = False

    if requested and selected is None:
        raise AccountSelectionError("La cuenta solicitada no está disponible como alternativa válida.")
    if selected:
        confidence = selected["estado_confianza"]
        trace = _trace_text(original, selected, cuit, criterion, confidence)
    elif requires:
        # Multiple existing accounts with no verifiable priority remain ambiguous
        # until an authorized user explicitly selects one. Never use CRM ordering.
        confidence = "RELACION_AMBIGUA"
        warnings.append(
            f"El CRM devolvio {len(found)} alternativas validas; requiere seleccion manual."
        )
        trace = None
    else:
        confidence = "SIN_RELACION"
        trace = None
    return {
        **relationship,
        "crm_modo": crm_mode,
        "cuentas_candidatas": candidates,
        "cuentas_encontradas_en_crm": found,
        "cantidad_evaluadas_crm": len(crm_results),
        "cuenta_seleccionada": selected,
        "criterio_seleccion": criterion,
        "estado_confianza": confidence,
        "confianza_label": _CONFIDENCE_LABELS[confidence],
        "confirmacion_fiscal": bool(selected and selected.get("confirmacion_fiscal")),
        "seleccion_origen": origin,
        "requiere_seleccion": requires,
        "bloqueado": selected is None,
        "advertencias": warnings,
        "trazabilidad_texto": trace,
    }