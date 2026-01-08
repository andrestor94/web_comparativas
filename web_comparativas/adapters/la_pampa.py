# -*- coding: utf-8 -*-
# Consolidador general (PDF): procesa TODOS los pares COMPARATIVA + PLIEGO en una carpeta
# Mantiene las mejoras: CGO, CANTIDAD, UNIDAD, DESCRIPCION completa (banda UNIDAD + banda DESCRIPCION)
# + NUEVO:
#   - Exporta también a formato de "normalizada estándar" basado en tu ejemplo normalized_85
#   - Agrega columna "Archivo Origen" en la salida estándar
#
# FIX 2025-12:
#   - Refuerzo fuerte del parser de COMPARATIVA:
#       * columnas SOLO si el token contiene "Alt"
#       * evita que números sueltos dentro del nombre del proveedor (ej: "20")
#         se interpreten como columnas o precios
#   - QC automático por TAG para alertas de:
#       * PUs sospechosos
#       * duplicados raros
#       * proveedores potencialmente truncados

import os
import re
from typing import Dict, List, Tuple, Optional, Any
from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
import numpy as np
from pathlib import Path
import zipfile
import shutil
import uuid

# ======================= CONFIG =======================
# Cambiá INPUT_DIR/OUTPUT_DIR a tu carpeta de trabajo
INPUT_DIR  = r"C:\Users\ANDRES.TORRES\Suizo Argentina S.A\grupo.Licitaciones - Licitaciones\N_Cuenta\8665"
OUTPUT_DIR = r"C:\Users\ANDRES.TORRES\Suizo Argentina S.A\grupo.Licitaciones - Licitaciones\Output\Consolidados\8665"
os.makedirs(OUTPUT_DIR, exist_ok=True)

MASTER_FILENAME   = "Consolidado_LaPampa.xlsx"
STANDARD_FILENAME = "Consolidado_LaPampa_ESTANDAR.xlsx"

# Estructura estándar EXACTA según tu archivo normalized_85 (1).xlsx
STANDARD_COLUMNS = [
    "Proveedor",
    "Renglón",
    "Alternativa",
    "Código",
    "Descripción",
    "Cantidad solicitada",
    "Unidad de medida",
    "Precio unitario",
    "Cantidad ofertada",
    "Total por renglón",
    "Especificación técnica",
    "Marca",
    "Posicion",
    "Archivo Origen",
]

# ======================= HELPERS =======================
FOOTER_NOISE_RE = re.compile(
    r'(Provincia\s+de\s+La\s+Pampa|DEPARTAMENTO\s+COMPRAS|CONTADURIA\s+GENERAL|'
    r'Firma\s+y\s+Sello\s+del\s+Proponente|Domicilio\s+Legal)',
    re.IGNORECASE
)

def _normalize_spaces(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s2 = re.sub(r'\s+', ' ', str(s)).strip()
    return s2 if s2 else None

def _is_nan(x) -> bool:
    try:
        return x is None or (isinstance(x, float) and np.isnan(x))
    except Exception:
        return False

def _round_half_up(val: float, decimals: int) -> float:
    q = Decimal(str(val)).quantize(Decimal('1.' + '0'*decimals), rounding=ROUND_HALF_UP)
    return float(q)

# ---------- números EU ----------
NUM_EU_RE = re.compile(r'^\d{1,3}(?:\.\d{3})*(?:,\d+)?$')

def _parse_number_eu_from_text(txt: str) -> Optional[float]:
    if txt is None:
        return None
    s = str(txt).strip()
    if s == "":
        return None
    s = (s.replace("ARS", "").replace("$", "").replace("€", "").replace("\u00A0", " ")
           .replace(" ", "").strip())
    s = re.sub(r'[^0-9\.,]', '', s)
    s = s.replace('.', '')
    if s.count(',') > 1:
        last = s.rfind(',')
        s = s[:last].replace(',', '') + s[last:]
    s = s.replace(',', '.')
    if s in {"", "."}:
        return None
    try:
        return float(Decimal(s))
    except Exception:
        m = re.search(r'\d+(?:\.\d+)?', s)
        return float(m.group(0)) if m else None

def _parse_number_cell(val: Any, *, decimals: Optional[int] = None, for_quantity: bool = False) -> Optional[float]:
    if _is_nan(val):
        return None
    try:
        num = float(val)
        if for_quantity and decimals is not None:
            return _round_half_up(num, decimals)
        return num
    except Exception:
        pass
    num = _parse_number_eu_from_text(str(val))
    if num is None:
        return None
    if for_quantity and decimals is not None:
        return _round_half_up(num, decimals)
    return num

# ---------- unidades ----------
_UNIT_CANON = {
    "UN": "UN", "UNI": "UN", "UNI.": "UN", "UN.": "UN", "U.": "UN", "UNIDAD": "UN",
    "ENV": "ENV", "ENV.": "ENV", "ENVASE": "ENV", "ENVASE.": "ENV", "ENVASES": "ENV", "ENVASES.": "ENV",
    "CAJA": "CAJA", "CAJA.": "CAJA",
    "SOBRE": "SOBRE", "SOBRE.": "SOBRE",
    "FRASCO": "FRASCO", "FRASCO.": "FRASCO",
    "BOLSA": "BOLSA", "BOLSA.": "BOLSA",
    "TUBO": "TUBO", "TUBO.": "TUBO",
    "PAR": "PAR", "PARES": "PAR",
    "KIT": "KIT",
    "ROLLO": "ROLLO", "ROLLO.": "ROLLO",
    "PAQUETE": "PAQUETE", "PAQ": "PAQUETE", "PAQ.": "PAQUETE",
    "JUEGO": "JUEGO", "JGO": "JUEGO",
    "JERINGA": "JERINGA",
    "BOTELLA": "BOTELLA",
    "POTE": "POTE",
    "AMPOLLA": "AMPOLLA", "AMP": "AMP", "AMP.": "AMP",
    "BIDON": "BIDON", "BIDÓN": "BIDON",
    "BOBINA": "BOBINA",
    "L": "L", "LITRO": "L", "ML": "ML", "CC": "CC",
    "KG": "KG", "G": "G",
    "M": "M", "METRO": "M",
    "CM": "CM", "MM": "MM"
}

def _canonical_unidad(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    t = _normalize_spaces(s)
    if not t:
        return None
    up = t.upper()
    for tok in re.split(r'[,:\s/|;·\-*]+', up):
        tok2 = tok.strip('.').strip()
        if not tok2:
            continue
        if tok2 in _UNIT_CANON:
            return _UNIT_CANON[tok2]
    return None

# Mapeo de visualización para la estructura estándar
_UNIT_DISPLAY = {
    "UN": "UNIDAD",
    "ENV": "ENVASE",
    "AMP": "AMPOLLA",
    "AMPOLLA": "AMPOLLA",
    "CAJA": "CAJA",
    "SOBRE": "SOBRE",
    "FRASCO": "FRASCO",
    "BOLSA": "BOLSA",
    "TUBO": "TUBO",
    "PAR": "PAR",
    "KIT": "KIT",
    "ROLLO": "ROLLO",
    "PAQUETE": "PAQUETE",
    "JUEGO": "JUEGO",
    "JERINGA": "JERINGA",
    "BOTELLA": "BOTELLA",
    "POTE": "POTE",
    "BIDON": "BIDON",
    "BOBINA": "BOBINA",
    "L": "L",
    "ML": "ML",
    "CC": "CC",
    "KG": "KG",
    "G": "G",
    "M": "M",
    "CM": "CM",
    "MM": "MM",
}

def _unit_display(u: Any) -> Optional[str]:
    if u is None or (isinstance(u, float) and np.isnan(u)):
        return None
    s = str(u).strip().upper()
    return _UNIT_DISPLAY.get(s, s)

# ---------- split desc/unidad SOLO si hay '$$' ----------
def _split_desc_and_unit_from_text(desc_raw: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not desc_raw:
        return (None, None)
    txt = _normalize_spaces(desc_raw) or ""
    if FOOTER_NOISE_RE.search(txt):
        return (None, None)
    t = re.sub(r'\$\s*\$', '$$', txt)
    if '$$' in t:
        left, right = t.split('$$', 1)
        desc = _normalize_spaces(left)
        munit = re.search(r'^([A-Za-zÁÉÍÓÚÑ./()%0-9 ]{1,20})', right.strip())
        unidad = _canonical_unidad(munit.group(1)) if munit else None
        return (desc, unidad)
    return (_normalize_spaces(txt), None)

# ---------- anclas por tríos numéricos ----------
def _detect_column_anchors(words: List[dict]) -> Tuple[float, float, float, float, float]:
    num_pat = re.compile(r'^\d{1,3}(?:\.\d{3})*(?:,\d+)?$')
    int_pat = re.compile(r'^\d{1,6}$')

    y_tol = 1.8
    def ybin(y): return round(y / y_tol) * y_tol
    lines: Dict[float, List[dict]] = {}
    for w in words or []:
        lines.setdefault(ybin(w['top']), []).append(w)

    xs_item, xs_cgo, xs_qty = [], [], []
    for _y, ws in lines.items():
        nums = [w for w in sorted(ws, key=lambda t: t['x0'])
                if num_pat.fullmatch(w['text']) or int_pat.fullmatch(w['text'])]
        if len(nums) < 3:
            continue
        if nums[0]['x0'] > 110:
            continue
        xs_item.append(nums[0]['x0'])
        xs_cgo.append(nums[1]['x0'])
        xs_qty.append(nums[2]['x0'])

    if xs_item and xs_cgo and xs_qty:
        x_item = float(np.median(xs_item))
        x_cgo  = float(np.median(xs_cgo))
        x_qty  = float(np.median(xs_qty))
        x_unit = x_qty + 40.0
        all_xs = [round(w['x0'], 1) for w in words]
        x_desc = float(np.percentile(all_xs, 80)) if all_xs else (x_unit + 120.0)
        return (x_item, x_cgo, x_qty, x_unit, x_desc)

    xs = sorted(set(round(w['x0'],1) for w in words))
    if xs:
        p = np.percentile(xs, [5, 20, 40, 60, 75])
        x_item, x_cgo, x_qty, x_unit, x_desc = float(p[0]), float(p[1]), float(p[2]), float(p[3]), float(p[4])
    else:
        x_item, x_cgo, x_qty, x_unit, x_desc = 56.0, 96.0, 146.0, 190.0, 270.0
    return (x_item, x_cgo, x_qty, x_unit, x_desc)

def _build_cuts_from_anchors(x_item: float, x_cgo: float, x_qty: float, x_unit: float, x_desc: float) -> List[float]:
    return [ (x_item+x_cgo)/2.0, (x_cgo+x_qty)/2.0, (x_qty+x_unit)/2.0, (x_unit+x_desc)/2.0 ]

def _assign_band(x: float, cuts: List[float]) -> int:
    if x < cuts[0]: return 0  # ITEM
    if x < cuts[1]: return 1  # CGO
    if x < cuts[2]: return 2  # CANTIDAD
    if x < cuts[3]: return 3  # UNIDAD (y comienzo de descripción)
    return 4                  # DESCRIPCION a la derecha

def _pick_qty_from_words(ws: List[dict], x_qty: float, x_unit: float) -> Optional[str]:
    cands = [w for w in ws if NUM_EU_RE.fullmatch(w['text'])
             and (x_qty - 40) <= w['x0'] <= (x_unit + 40)]
    if not cands:
        return None
    mid = (x_qty + x_unit)/2.0
    w = min(cands, key=lambda t: abs(t['x0']-mid))
    return w['text']

def _extract_unit_and_b3desc(b3_words: List[dict]) -> Tuple[Optional[str], Optional[str]]:
    """
    En la banda 3: detecta UNIDAD y devuelve todo lo que quede a la derecha
    como prefijo de DESCRIPCION. Si no hay unidad, toda la banda 3 va a DESCRIPCION.
    """
    if not b3_words:
        return (None, None)
    ws = sorted(b3_words, key=lambda t: t['x0'])
    unit = None
    unit_idx = None
    for i, w in enumerate(ws):
        cand = _canonical_unidad(w['text'])
        if cand:
            unit = cand
            unit_idx = i
            break
    if unit is not None and unit_idx is not None:
        rest = [w['text'] for w in ws[unit_idx+1:]]
        extra = _normalize_spaces(" ".join(rest)) if rest else None
        return (unit, extra)
    extra = _normalize_spaces(" ".join(w['text'] for w in ws))
    return (None, extra if extra else None)

# ======================= Tagging & expediente =======================
def _extract_expediente_from_pdf(path_pdf: str) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(path_pdf) as pdf:
            for page in pdf.pages[:3]:
                txt = page.extract_text() or ""
                m = re.search(r'Expediente\s*N[º°]?:?\s*([0-9A-Za-z\/\-.]+)', txt, re.IGNORECASE)
                if m:
                    return m.group(1).strip()
    except Exception:
        pass
    # fallback: por nombre de archivo
    base = os.path.basename(path_pdf)
    m2 = re.search(r'(\d{3,6})[-_/](\d{2,4})', base)
    if m2:
        return f"{m2.group(1)}/{m2.group(2)}"
    return "SIN_EXPEDIENTE"

def _tag_expediente(expediente: str) -> str:
    m = re.search(r'([0-9]{3,6}).*?([0-9]{2,4})', expediente or "")
    if not m:
        solo_dig = re.sub(r'\D+', '', expediente or "")
        return solo_dig[-6:] if solo_dig else "SIN_TAG"
    num = int(m.group(1)); year = int(m.group(2))
    if year >= 100: year = year % 100
    return f"{num}-{year:02d}"

# ======================= Lectores (PDF) =======================
def leer_pliego_pdf(path_pdf: str) -> pd.DataFrame:
    """
    - Detecta columnas por tríos numéricos (ITEM, CGO, CANT).
    - UNIDAD: banda 3 (y lo sobrante se concatena como prefijo de DESCRIPCION),
      + soporte '$$' con lookahead.
    """
    try:
        import pdfplumber
    except Exception:
        raise RuntimeError("Instalá 'pdfplumber' para leer pliegos PDF.")

    regs: List[Dict[str, Any]] = []
    num_pat = re.compile(r'^\d{1,3}(?:\.\d{3})*(?:,\d+)?$')
    int_pat = re.compile(r'^\d{1,6}$')

    with pdfplumber.open(path_pdf) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            if not words:
                continue

            x_item, x_cgo, x_qty, x_unit, x_desc = _detect_column_anchors(words)
            cuts = _build_cuts_from_anchors(x_item, x_cgo, x_qty, x_unit, x_desc)

            y_tol = 1.8
            def ybin(y): return round(y / y_tol) * y_tol
            line_map: Dict[float, List[dict]] = {}
            for w in words:
                line_map.setdefault(ybin(w['top']), []).append(w)

            last_row: Optional[Dict[str, Any]] = None
            expect_unit_lines = 0

            for y in sorted(line_map.keys()):
                ws = sorted(line_map[y], key=lambda w: w['x0'])
                band_words = {0: [], 1: [], 2: [], 3: [], 4: []}
                for w in ws:
                    band_words[_assign_band(w['x0'], cuts)].append(w)

                def join_text(idx):
                    return _normalize_spaces(" ".join(t['text'] for t in band_words[idx]).strip()) if band_words[idx] else None

                item_txt = join_text(0)
                cgo_txt  = join_text(1)
                qty_txt0 = join_text(2)
                unit_from_b3, desc_prefix_b3 = _extract_unit_and_b3desc(band_words[3])
                desc_txt_b4 = join_text(4)

                # filtrar ruido
                if cgo_txt and FOOTER_NOISE_RE.search(cgo_txt): cgo_txt = None
                if qty_txt0 and FOOTER_NOISE_RE.search(qty_txt0): qty_txt0 = None
                if desc_txt_b4 and FOOTER_NOISE_RE.search(desc_txt_b4): desc_txt_b4 = None

                nums_on_line = [w for w in ws if NUM_EU_RE.fullmatch(w['text']) or re.fullmatch(r'\d{1,6}', w['text'])]
                nums_on_line.sort(key=lambda t: t['x0'])

                # nuevo ítem
                is_item_line = any(re.fullmatch(r'\d{1,6}', w['text']) and abs(w['x0'] - x_item) <= 20 for w in nums_on_line)
                if is_item_line:
                    if last_row:
                        regs.append(last_row); last_row = None; expect_unit_lines = 0

                    item_candidates = [w for w in nums_on_line if re.fullmatch(r'\d{1,6}', w['text']) and abs(w['x0'] - x_item) <= 20]
                    if not item_candidates:
                        continue
                    item_val = int(item_candidates[0]['text'])

                    # CGO
                    cgo_val = None
                    if len(nums_on_line) >= 2:
                        cgo_val = nums_on_line[1]['text']
                    if cgo_val is not None and not (NUM_EU_RE.fullmatch(cgo_val) or re.fullmatch(r'\d{1,6}', cgo_val)):
                        cgo_val = None
                    if cgo_val is None:
                        near_cgo = [w for w in nums_on_line if abs(w['x0'] - x_cgo) <= 20]
                        if near_cgo:
                            cgo_val = near_cgo[0]['text']
                    cgo_val = _normalize_spaces(cgo_val)

                    # CANTIDAD
                    qty_val = None
                    if len(nums_on_line) >= 3:
                        qty_val = _parse_number_cell(nums_on_line[2]['text'], decimals=2, for_quantity=True)
                    if qty_val is None:
                        qty_txt = qty_txt0 if (qty_txt0 and NUM_EU_RE.fullmatch(qty_txt0)) else _pick_qty_from_words(ws, x_qty, x_unit)
                        qty_val = _parse_number_cell(qty_txt, decimals=2, for_quantity=True)

                    # DESCRIPCION (prefijo banda 3 + banda 4)
                    desc_line = " ".join([t for t in [desc_prefix_b3, desc_txt_b4] if t])
                    desc_clean, unit_from_desc = _split_desc_and_unit_from_text(desc_line or "")

                    # UNIDAD
                    unidad_val = unit_from_b3 or (_canonical_unidad(unit_from_desc) if unit_from_desc else None)
                    if (desc_line and '$' in desc_line) and not unidad_val:
                        expect_unit_lines = 3

                    last_row = {
                        "ITEM": item_val,
                        "CGO": cgo_val,
                        "CANTIDAD": qty_val,
                        "DESCRIPCION": desc_clean,
                        "UNIDAD": unidad_val
                    }
                    continue

                # continuación
                if last_row:
                    if (not last_row.get('UNIDAD')) and unit_from_b3:
                        last_row['UNIDAD'] = unit_from_b3

                    if desc_prefix_b3 or desc_txt_b4:
                        prev = last_row.get('DESCRIPCION') or ""
                        joined_desc = _normalize_spaces((" ".join([prev, desc_prefix_b3 or "", desc_txt_b4 or ""])).strip())
                        d2, u2 = _split_desc_and_unit_from_text(joined_desc)
                        last_row['DESCRIPCION'] = d2
                        if (not last_row.get('UNIDAD')) and u2:
                            last_row['UNIDAD'] = _canonical_unidad(u2)
                        if (desc_prefix_b3 and '$' in desc_prefix_b3) or (desc_txt_b4 and '$' in desc_txt_b4):
                            if not last_row.get('UNIDAD'):
                                expect_unit_lines = max(expect_unit_lines, 3)

                    if last_row.get('CANTIDAD') is None:
                        qty_txt = qty_txt0 if (qty_txt0 and NUM_EU_RE.fullmatch(qty_txt0)) else _pick_qty_from_words(ws, x_qty, x_unit)
                        q2 = _parse_number_cell(qty_txt, decimals=2, for_quantity=True)
                        if q2 is not None:
                            last_row['CANTIDAD'] = q2

                    if expect_unit_lines > 0 and not last_row.get('UNIDAD'):
                        line_text_all = _normalize_spaces(" ".join(t['text'] for t in ws))
                        cand2 = _canonical_unidad(line_text_all or "")
                        if cand2:
                            last_row['UNIDAD'] = cand2
                        expect_unit_lines -= 1

            if last_row:
                regs.append(last_row); last_row = None

    df = pd.DataFrame(regs)
    if df.empty:
        return pd.DataFrame(columns=['ITEM','CGO','CANTIDAD','DESCRIPCION','UNIDAD'])

    # consolidar por ITEM
    def _modo_str(series: pd.Series) -> Optional[str]:
        s = series.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty: return None
        return s.value_counts(dropna=True).index[0]

    def _longest_str(series: pd.Series) -> Optional[str]:
        s = series.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty: return None
        return max(s, key=len)

    def _modo_num(series: pd.Series) -> Optional[float]:
        s = series.dropna().astype(float)
        if s.empty: return None
        vc = s.round(6).value_counts()
        cands = vc[vc == vc.max()].index.tolist()
        return float(min(cands))

    out = df.groupby('ITEM', as_index=False).agg({
        'CGO': _modo_str,
        'UNIDAD': _modo_str,
        'DESCRIPCION': _longest_str,
        'CANTIDAD': _modo_num
    })
    with pd.option_context('mode.chained_assignment', None):
        out['CANTIDAD'] = out['CANTIDAD'].apply(lambda x: _round_half_up(float(x), 2) if x is not None else None)
        out['CGO'] = out['CGO'].apply(_normalize_spaces)
        out['UNIDAD'] = out['UNIDAD'].apply(_canonical_unidad)
        out['DESCRIPCION'] = out['DESCRIPCION'].apply(_normalize_spaces)

    return out[['ITEM', 'CGO', 'CANTIDAD', 'DESCRIPCION', 'UNIDAD']]

# ---------- comparativa ----------
def _parse_item_alt_token(txt: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Se mantiene para compatibilidad con otras partes del script,
    pero el parser robusto de comparativas usa un STRICT interno.
    """
    if not txt:
        return (None, None)
    s = str(txt).strip()
    m = re.search(r'(\d+)\s*(?:-\s*\d+)?\s*-\s*Alt\.?\s*(\d+)', s, flags=re.IGNORECASE)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    m2 = re.search(r'(\d+)\s+Alt\.?\s*(\d+)', s, flags=re.IGNORECASE)
    if m2:
        return (int(m2.group(1)), int(m2.group(2)))
    if re.fullmatch(r'\d+', s):
        return (int(s), 0)
    return (None, None)

def leer_comparativa_pdf(path_pdf: str) -> pd.DataFrame:
    """
    Parser posicional robusto:
      - Detecta columnas SOLO con tokens que incluyen 'Alt'
      - PROVEEDOR: texto a la izquierda de la primer columna real
      - PU: números EU alineados a cada columna
      - ignora TOTAL/PROMEDIO
      - evita que números sueltos del proveedor (ej: '20') se interpreten
        como columnas o precios.
    """
    try:
        import pdfplumber
    except Exception:
        raise RuntimeError("Instalá 'pdfplumber' para leer comparativas PDF.")

    registros: List[Dict[str, Any]] = []
    num_pat = re.compile(r'^\d{1,3}(?:\.\d{3})*(?:,\d+)?$')

    # ---- parser ESTRICTO para columnas ----
    def _parse_item_alt_token_strict(txt: str) -> Tuple[Optional[int], Optional[int]]:
        if not txt:
            return (None, None)
        s = str(txt).strip()
        m = re.search(r'(\d+)\s*(?:-\s*\d+)?\s*-\s*Alt\.?\s*(\d+)', s, flags=re.IGNORECASE)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        m2 = re.search(r'(\d+)\s+Alt\.?\s*(\d+)', s, flags=re.IGNORECASE)
        if m2:
            return (int(m2.group(1)), int(m2.group(2)))
        return (None, None)

    with pdfplumber.open(path_pdf) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            if not words:
                continue

            # 1) Detectar columnas SOLO con 'Alt'
            code_words = []
            for w in words:
                it, al = _parse_item_alt_token_strict(w['text'])
                if it is not None:
                    code_words.append(w)

            if not code_words:
                continue

            xs = sorted([w['x0'] for w in code_words])
            cols_x = []
            cluster = [xs[0]]
            for x in xs[1:]:
                if abs(x - float(np.mean(cluster))) <= 12:
                    cluster.append(x)
                else:
                    cols_x.append(float(np.mean(cluster)))
                    cluster = [x]
            cols_x.append(float(np.mean(cluster)))
            code_cols = sorted(set(cols_x))

            # mapa idx -> (item, alt, rawtxt)
            colcodes: Dict[int, Tuple[int, int, str]] = {}
            for w in code_words:
                idx = int(np.argmin([abs(w['x0'] - cx) for cx in code_cols]))
                it, al = _parse_item_alt_token_strict(w['text'])
                if it is not None and idx not in colcodes:
                    colcodes[idx] = (int(it), int(al), w['text'])

            if not colcodes:
                continue

            min_code_x = min(code_cols)
            x_thresh = min_code_x - 12.0
            x_left_limit = x_thresh + 25.0  # zona proveedor razonable

            # 2) Agrupar por líneas
            y_tol = 1.8
            def ybin(y): return round(y / y_tol) * y_tol

            line_map: Dict[float, Dict[str, List[dict]]] = {}
            for w in words:
                y = ybin(w['top'])
                obj = line_map.setdefault(y, {'left': [], 'nums': [], 'all': []})
                obj['all'].append(w)

                # palabras zona izquierda (proveedor)
                if w['x0'] < x_left_limit:
                    obj['left'].append(w)

                # números candidatos a precios: solo si están en zona columnas
                if num_pat.fullmatch(w['text']) and (w['x0'] >= (min_code_x - 5)):
                    obj['nums'].append(w)

            ys = sorted(line_map.keys())

            # 3) Join inteligente del proveedor
            def _join_provider(ws: List[dict]) -> Optional[str]:
                if not ws:
                    return None
                ws_sorted = sorted(ws, key=lambda t: t['x0'])
                has_letters = any(re.search(r'[A-Za-zÁÉÍÓÚÑ]', t['text']) for t in ws_sorted)

                keep = []
                for t in ws_sorted:
                    tx = str(t['text']).strip()
                    if not tx:
                        continue
                    if re.search(r'[A-Za-zÁÉÍÓÚÑ]', tx):
                        keep.append(tx)
                        continue
                    # permitir números cortos si hay letras en la línea
                    if has_letters and re.fullmatch(r'\d{1,3}', tx):
                        keep.append(tx)
                        continue

                return _normalize_spaces(" ".join(keep))

            def _get_provider(i: int) -> Optional[str]:
                if line_map[ys[i]]['left']:
                    p = _join_provider(line_map[ys[i]]['left'])
                    if p:
                        return p

                for k in range(1, 6):
                    j = i - k
                    if j < 0: break
                    if ys[i] - ys[j] > 60: break
                    if line_map[ys[j]]['left']:
                        p = _join_provider(line_map[ys[j]]['left'])
                        if p:
                            return p

                if i + 1 < len(ys) and (ys[i + 1] - ys[i] <= 24.0) and line_map[ys[i + 1]]['left']:
                    p = _join_provider(line_map[ys[i + 1]]['left'])
                    if p:
                        return p

                return None

            # 4) Construir registros
            for ii, y in enumerate(ys):
                row = line_map[y]
                if not row['nums']:
                    continue

                prov = _get_provider(ii)
                if not prov:
                    continue

                up = prov.upper()
                if up.startswith("TOTAL") or up.startswith("PROMEDIO"):
                    continue
                if up.startswith("PROVEEDOR"):
                    continue

                for numw in sorted(row['nums'], key=lambda t: t['x0']):
                    idx = int(np.argmin([abs(numw['x0'] - cx) for cx in code_cols]))
                    if idx not in colcodes:
                        continue

                    # filtro extra: el número debe estar razonablemente cerca de su columna
                    if abs(numw['x0'] - code_cols[idx]) > 35:
                        continue

                    item, alt, rawtxt = colcodes[idx]
                    pu = _parse_number_cell(numw['text'])
                    if pu is None:
                        continue

                    registros.append({
                        "PROVEEDOR": prov.strip(),
                        "ITEM": int(item),
                        "Alt": int(alt),
                        "PU": float(pu),
                        "ItemCode": rawtxt or f"{item}-0-Alt.{alt}"
                    })

    cols = ["PROVEEDOR", "ITEM", "Alt", "PU", "ItemCode"]
    return pd.DataFrame(registros, columns=cols) if registros else pd.DataFrame(columns=cols)

# ======== QC COMPARATIVA ========
def _qc_comparativa_df(df_comp: pd.DataFrame, tag_name: str) -> None:
    """
    QC rápido y útil en consola para detectar:
      - PUs sospechosos bajos (default <= 30)
      - duplicados raros por (PROVEEDOR, ITEM, Alt, PU)
      - proveedores potencialmente truncados (heurística simple)
    """
    try:
        if df_comp is None or df_comp.empty:
            print(f"[QC {tag_name} | COMPARATIVA] Sin filas.")
            return

        n = len(df_comp)
        n_prov = df_comp["PROVEEDOR"].nunique(dropna=True) if "PROVEEDOR" in df_comp.columns else 0
        print(f"[QC {tag_name} | COMPARATIVA] Filas: {n} | Proveedores únicos: {n_prov}")

        # 1) PUs sospechosos muy bajos
        if "PU" in df_comp.columns:
            low = df_comp[df_comp["PU"].notna() & (df_comp["PU"] <= 30)]
            if not low.empty:
                cnt_low = len(low)
                print(f"[ALERTA {tag_name}] PUs <= 30 detectados: {cnt_low}")
                print(low[["PROVEEDOR", "ITEM", "Alt", "PU"]].head(20).to_string(index=False))

        # 2) Duplicados raros exactos
        key_cols = [c for c in ["PROVEEDOR", "ITEM", "Alt", "PU"] if c in df_comp.columns]
        if len(key_cols) == 4:
            dup = (df_comp.groupby(key_cols)
                         .size()
                         .reset_index(name="count")
                         .sort_values("count", ascending=False))
            dup = dup[dup["count"] >= 5]
            if not dup.empty:
                print(f"[ALERTA {tag_name}] Duplicados fuertes (>=5) en (PROVEEDOR, ITEM, Alt, PU):")
                print(dup.head(15).to_string(index=False))

        # 3) Heurística de proveedor truncado
        if "PROVEEDOR" in df_comp.columns:
            provs = (df_comp["PROVEEDOR"].dropna().astype(str).map(_normalize_spaces)).dropna().unique().tolist()
            suspicious = []
            for p in provs:
                toks = p.split()
                if len(p) < 12:
                    suspicious.append(p)
                    continue
                if len(toks) <= 2:
                    suspicious.append(p)
                    continue
                # termina en preposición común -> probable truncado
                if toks[-1].upper() in {"DE", "DEL", "LA", "LAS", "LOS", "Y"}:
                    suspicious.append(p)

            if suspicious:
                suspicious = sorted(set(suspicious), key=len)
                print(f"[ALERTA {tag_name}] Proveedores potencialmente truncados:")
                for p in suspicious[:20]:
                    print(f"  - {p}")

    except Exception:
        # QC no debe cortar el pipeline
        pass

# ======== util Excel ========
def _choose_writer_engine() -> str:
    try:
        import xlsxwriter  # noqa: F401
        return "xlsxwriter"
    except Exception:
        return "openpyxl"

def _format_excel(path_xlsx: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(path_xlsx)
        for ws in wb.worksheets:
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="CDE7FF")
            thin = Side(border_style="thin", color="999999")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
            ws.freeze_panes = ws["A2"]
        wb.save(path_xlsx)
    except Exception:
        pass

# ======== detectores de tipo ========
def _looks_like_comparativa_pdf(path: str) -> bool:
    """Heurístico: nombre contiene 'comparativa' o en la primer página aparecen códigos 'Alt.'"""
    name_ok = 'comparativa' in os.path.basename(path).lower()
    if name_ok:
        return True
    # inspección rápida
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            page = pdf.pages[0]
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False) or []
            return any(re.search(r'\bAlt\.?\s*\d+', w['text'], flags=re.IGNORECASE) for w in words)
    except Exception:
        return False

def _looks_like_pliego_pdf(path: str) -> bool:
    """Heurístico: nombre contiene '-lp'/'pliego' o en el texto aparecen cabeceras de ítems."""
    name = os.path.basename(path).lower()
    if ('-lp' in name) or ('pliego' in name):
        return True
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            txt = (pdf.pages[0].extract_text() or "").upper()
            return any(t in txt for t in ["ITEM", "ÍTEM", "RENGL", "DESCRIP", "UNIDAD"])
    except Exception:
        return False

# ======== pairing por TAG ========
def _find_tag_in_filename(path: str) -> Optional[str]:
    base = os.path.basename(path)
    m = re.search(r'(\d{3,6})[-_/](\d{2})', base)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2)):02d}"
    return None

def _pair_pliego_for_tag(tag: str, pliego_paths: List[str]) -> Optional[str]:
    # 1) por nombre
    for p in pliego_paths:
        if tag in os.path.basename(p):
            return p
    # 2) por expediente dentro del PDF
    for p in pliego_paths:
        exp = _extract_expediente_from_pdf(p)
        if exp and _tag_expediente(exp) == tag:
            return p
    return None

# ======== NUEVO: builder de salida estándar ========
def _build_standard_from_master(master_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convierte el dataframe Master actual al formato estándar
    basado en tu ejemplo normalized_85 (1).xlsx
    y agrega 'Archivo Origen'.
    """
    if master_df is None or master_df.empty:
        return pd.DataFrame(columns=STANDARD_COLUMNS)

    out = pd.DataFrame(index=master_df.index)

    out["Proveedor"] = master_df.get("PROVEEDOR")
    out["Renglón"] = master_df.get("ITEM")
    out["Alternativa"] = master_df.get("Alt")
    out["Código"] = master_df.get("CGO")
    out["Descripción"] = master_df.get("DESCRIPCION")
    out["Cantidad solicitada"] = master_df.get("CANTIDAD")

    if "UNIDAD" in master_df.columns:
        out["Unidad de medida"] = master_df["UNIDAD"].apply(_unit_display)
    else:
        out["Unidad de medida"] = None

    out["Precio unitario"] = master_df.get("PU")

    # No existe en tu consolidado actual: lo dejamos vacío
    out["Cantidad ofertada"] = pd.Series([None] * len(master_df), index=master_df.index)

    out["Total por renglón"] = master_df.get("TOTAL")

    # No existen en tu consolidado actual: vacíos
    out["Especificación técnica"] = pd.Series([None] * len(master_df), index=master_df.index)
    out["Marca"] = pd.Series([None] * len(master_df), index=master_df.index)

    out["Posicion"] = master_df.get("Posicion")

    # Archivo Origen (preferimos comparativa)
    if "ArchivoComparativa" in master_df.columns:
        out["Archivo Origen"] = master_df["ArchivoComparativa"]
    elif "ArchivoPliego" in master_df.columns:
        out["Archivo Origen"] = master_df["ArchivoPliego"]
    else:
        out["Archivo Origen"] = None

    # Asegurar orden exacto
    for c in STANDARD_COLUMNS:
        if c not in out.columns:
            out[c] = None

    return out[STANDARD_COLUMNS]

# ======== pipeline general ========
def run_all_pipelines(input_dir: str = INPUT_DIR, output_dir: str = OUTPUT_DIR, master_filename: str = MASTER_FILENAME):
    pdfs = [os.path.join(input_dir, f) for f in os.listdir(input_dir) if f.lower().endswith(".pdf")]
    if not pdfs:
        print("No se encontraron PDFs en la carpeta.")
        return

    comparativas = [p for p in pdfs if _looks_like_comparativa_pdf(p)]
    pliegos     = [p for p in pdfs if _looks_like_pliego_pdf(p)]

    if not comparativas:
        print("No se detectaron COMPARATIVAS (PDF).")
        return
    if not pliegos:
        print("No se detectaron PLIEGOS (PDF).")
        return

    master_rows: List[pd.DataFrame] = []
    por_tag: Dict[str, pd.DataFrame] = {}
    procesados = 0

    for comp in sorted(comparativas):
        expediente = _extract_expediente_from_pdf(comp)
        tag_name = _find_tag_in_filename(comp) or _tag_expediente(expediente)
        if not tag_name or tag_name == "SIN_TAG":
            print(f"[AVISO] No se pudo derivar TAG para comparativa: {os.path.basename(comp)}")
            continue

        pliego_file = _pair_pliego_for_tag(tag_name, pliegos)
        if not pliego_file:
            print(f"[AVISO] No se encontró pliego para TAG {tag_name} (comparativa: {os.path.basename(comp)}). Se omite.")
            continue

        print(f"\n==> Pareja detectada (TAG {tag_name}):")
        print(f"   COMPARATIVA: {os.path.basename(comp)}")
        print(f"   PLIEGO     : {os.path.basename(pliego_file)}")
        print(f"   Expediente : {expediente}")

        # Leer
        df_pl   = leer_pliego_pdf(pliego_file)
        df_comp = leer_comparativa_pdf(comp)

        # QC comparativa (antes del merge)
        _qc_comparativa_df(df_comp, tag_name)

        # Merge
        df = df_comp.merge(df_pl[['ITEM','CGO','CANTIDAD','DESCRIPCION','UNIDAD']], on="ITEM", how="left")

        # Atributos fijos
        df["EXPEDIENTE"] = expediente
        df["Tag"] = tag_name
        df["ArchivoComparativa"] = os.path.basename(comp)
        df["ArchivoPliego"] = os.path.basename(pliego_file)

        # TOTAL y ranking
        df["TOTAL"] = df.apply(
            lambda r: (r["PU"] * r["CANTIDAD"]) if pd.notna(r["PU"]) and pd.notna(r["CANTIDAD"]) else None,
            axis=1
        )
        if not df.empty:
            ranked = df.groupby(["ITEM","Alt"])['PU'].rank(method="dense", ascending=True)
            df["Posicion"] = ranked.mask(df["PU"].isna()).astype("Int64")
        else:
            df["Posicion"] = pd.Series(dtype="Int64")

        # Orden columnas Master
        cols = [
            "Tag", "EXPEDIENTE",
            "ITEM", "Alt", "CGO", "CANTIDAD", "UNIDAD", "DESCRIPCION",
            "PROVEEDOR", "PU", "TOTAL", "Posicion",
            "ItemCode", "ArchivoComparativa", "ArchivoPliego"
        ]
        for c in cols:
            if c not in df.columns:
                df[c] = None
        df = df[cols].sort_values(["Tag","ITEM","Alt","Posicion","PROVEEDOR"], na_position="last").reset_index(drop=True)

        # QC por TAG (pliego+merge)
        try:
            n = len(df)
            n_qty_nan  = int(df['CANTIDAD'].isna().sum())
            n_unit_nan = int(df['UNIDAD'].isna().sum())
            n_cgo_nan  = int(df['CGO'].isna().sum())
            print(f"[QC {tag_name}] Filas: {n} | CANTIDAD vacías: {n_qty_nan} | UNIDAD vacías: {n_unit_nan} | CGO vacíos: {n_cgo_nan}")
        except Exception:
            pass

        master_rows.append(df.copy())
        por_tag[tag_name] = df.copy()
        procesados += 1

    if not master_rows:
        print("No se generó contenido para el Master.")
        return

    master_df = pd.concat(master_rows, ignore_index=True)

    master_path = os.path.join(output_dir, master_filename)

    # Escribimos Master + una hoja por TAG
    with pd.ExcelWriter(master_path, engine=_choose_writer_engine()) as xw:
        master_df.to_excel(xw, sheet_name="Master", index=False)
        usados = set()
        for tag, dft in por_tag.items():
            name = tag[:31]
            base = name
            k = 1
            while name in usados:
                suf = f"_{k}"
                name = (base[:(31 - len(suf))] + suf)
                k += 1
            usados.add(name)
            dft.to_excel(xw, sheet_name=name, index=False)

    _format_excel(master_path)

    # ===== NUEVO: export estándar =====
    standard_path = os.path.join(output_dir, STANDARD_FILENAME)

    std_master = _build_standard_from_master(master_df)

    std_por_tag: Dict[str, pd.DataFrame] = {}
    for tag, dft in por_tag.items():
        std_por_tag[tag] = _build_standard_from_master(dft)

    with pd.ExcelWriter(standard_path, engine=_choose_writer_engine()) as xw:
        std_master.to_excel(xw, sheet_name="Estandar", index=False)

        usados = set()
        for tag, dft in std_por_tag.items():
            name = tag[:31]
            base = name
            k = 1
            while name in usados:
                suf = f"_{k}"
                name = (base[:(31 - len(suf))] + suf)
                k += 1
            usados.add(name)
            dft.to_excel(xw, sheet_name=name, index=False)

    _format_excel(standard_path)

    print(f"\n[OK] Generado Master: {master_path}")
    print(f"[OK] Generado Estándar: {standard_path}")
    print(f"¡Listo! Pares procesados: {procesados}")

# ======== ADAPTER ENTRY POINT for Web App ========
def normalize_la_pampa(input_path: Path, metadata: dict, out_dir: Path) -> dict:
    """
    Adapter function for the web application.
    Recibe un ZIP con (al menos) 1 Comparativa y 1 Pliego.
    Retorna {'df': standard_df, 'summary': ...}
    """
    import sys
    
    print(f"--- START NORMALIZE_LA_PAMPA ---")
    print(f"Input: {input_path}")
    
    # 1. Unzip
    extract_dir = out_dir / f"extract_{uuid.uuid4().hex[:8]}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        with zipfile.ZipFile(input_path, 'r') as zf:
            zf.extractall(extract_dir)
            
        # 2. Identify files
        all_files = [str(p) for p in extract_dir.glob("**/*") if p.is_file() and p.suffix.lower() == ".pdf"]
        print(f"Extracted files: {all_files}")
        
        comparativas = [p for p in all_files if _looks_like_comparativa_pdf(p)]
        pliegos = [p for p in all_files if _looks_like_pliego_pdf(p)]
        
        print(f"Identified Comp: {comparativas}")
        print(f"Identified Plie: {pliegos}")
        
        # Fallback Logic
        if not comparativas and len(all_files) == 2:
            # Si hay 2 y uno es pliego, el otro es comparativa
            leftover = [p for p in all_files if p not in pliegos]
            if len(leftover) == 1:
                comparativas = leftover
                print(f"Fallback: using {comparativas[0]} as Comparativa")

        if not pliegos and len(all_files) == 2:
            # Si hay 2 y uno es comparativa, el otro es pliego
            leftover = [p for p in all_files if p not in comparativas]
            if len(leftover) == 1:
                pliegos = leftover
                print(f"Fallback: using {pliegos[0]} as Pliego")
        
        # Force fallback if still ambiguous but exactly 2 files
        if (not comparativas or not pliegos) and len(all_files) == 2:
             # Assume larger/smaller or sort by name? 
             # Let's try heuristic: "Pliego" usually has "Pliego" or "LP" but if user renamed it...
             # We just take the one that ISN'T the other if we found one.
             # If we found NEITHER, we are in trouble.
             pass

        if not comparativas:
            raise ValueError("No se detectó archivo de Comparativa (PDF) en el ZIP.")
        if not pliegos:
            raise ValueError("No se detectó archivo de Pliego (PDF) en el ZIP.")

        # 3. Process pairs
        master_rows = []
        
        # Caso simple: 1 vs 1 (Most common in Web App)
        if len(comparativas) == 1 and len(pliegos) == 1:
            comp_path = comparativas[0]
            pliego_path = pliegos[0]
            tag_name = "SINGLE_PAIR"
            expediente = _extract_expediente_from_pdf(comp_path)
            
            print(f"Processing SINGLE PAIR: Comp={Path(comp_path).name}, Pliego={Path(pliego_path).name}")
            
            df_pl = leer_pliego_pdf(pliego_path)
            df_comp = leer_comparativa_pdf(comp_path)
            
            print(f"DF Pliego shape: {df_pl.shape}")
            print(f"DF Comp shape: {df_comp.shape}")
            if not df_pl.empty:
                print(f"DF Pliego Head:\n{df_pl.head(2).to_string()}")
            else:
                print("WARNING: DF Pliego IS EMPTY!")

            # --- MERGE ROBUSTNESS ---
            # Ensure ITEM is int
            if not df_comp.empty and "ITEM" in df_comp.columns:
                 df_comp["ITEM"] = pd.to_numeric(df_comp["ITEM"], errors='coerce').fillna(0).astype(int)
            
            if not df_pl.empty and "ITEM" in df_pl.columns:
                 df_pl["ITEM"] = pd.to_numeric(df_pl["ITEM"], errors='coerce').fillna(0).astype(int)

            # Merge
            df = df_comp.merge(df_pl[['ITEM','CGO','CANTIDAD','DESCRIPCION','UNIDAD']], on="ITEM", how="left")
            print(f"Merged DF shape: {df.shape}")
            
            # Helper to fill NaNs if merge failed (debug check)
            if df["DESCRIPCION"].isna().all() and not df_pl.empty:
                print("WARNING: Merge resulted in all NaNs for Pliego columns. Check ITEM matching.")
                print(f"Comp Items: {df_comp['ITEM'].unique()[:10]}")
                print(f"Pliego Items: {df_pl['ITEM'].unique()[:10]}")
                
            df["EXPEDIENTE"] = expediente
            df["Tag"] = tag_name
            df["ArchivoComparativa"] = os.path.basename(comp_path)
            df["ArchivoPliego"] = os.path.basename(pliego_path)
            
            df["TOTAL"] = df.apply(
                lambda r: (r["PU"] * r["CANTIDAD"]) if pd.notna(r["PU"]) and pd.notna(r["CANTIDAD"]) else None,
                axis=1
            )
            
            if not df.empty:
                ranked = df.groupby(["ITEM","Alt"])['PU'].rank(method="dense", ascending=True)
                df["Posicion"] = ranked.mask(df["PU"].isna()).astype("Int64")
            else:
                df["Posicion"] = pd.Series(dtype="Int64")
                
            master_rows.append(df)
            
        else:
            # Caso complejo: varios archivos
            print(f"Processing MULTIPLE PAIRS logic.")
            for comp in sorted(comparativas):
                exp = _extract_expediente_from_pdf(comp)
                tag = _find_tag_in_filename(comp) or _tag_expediente(exp)
                pliego = _pair_pliego_for_tag(tag, pliegos)
                
                if pliego:
                    print(f"Matched Tag {tag}: {Path(comp).name} <-> {Path(pliego).name}")
                    df_pl = leer_pliego_pdf(pliego)
                    df_comp = leer_comparativa_pdf(comp)
                    
                    if not df_comp.empty and "ITEM" in df_comp.columns:
                         df_comp["ITEM"] = pd.to_numeric(df_comp["ITEM"], errors='coerce').fillna(0).astype(int)
                    if not df_pl.empty and "ITEM" in df_pl.columns:
                         df_pl["ITEM"] = pd.to_numeric(df_pl["ITEM"], errors='coerce').fillna(0).astype(int)

                    df = df_comp.merge(df_pl[['ITEM','CGO','CANTIDAD','DESCRIPCION','UNIDAD']], on="ITEM", how="left")
                    df["EXPEDIENTE"] = exp
                    df["Tag"] = tag
                    df["ArchivoComparativa"] = os.path.basename(comp)
                    df["ArchivoPliego"] = os.path.basename(pliego)
                    
                    df["TOTAL"] = df.apply(
                        lambda r: (r["PU"] * r["CANTIDAD"]) if pd.notna(r["PU"]) and pd.notna(r["CANTIDAD"]) else None,
                        axis=1
                    )
                    if not df.empty:
                        ranked = df.groupby(["ITEM","Alt"])['PU'].rank(method="dense", ascending=True)
                        df["Posicion"] = ranked.mask(df["PU"].isna()).astype("Int64")
                    else:
                        df["Posicion"] = pd.Series(dtype="Int64")
                    
                    master_rows.append(df)
                else:
                    print(f"WARNING: No matching Pliego for {Path(comp).name} (Tag: {tag})")
        
        if not master_rows:
             raise ValueError("No se pudieron procesar datos (posiblemente no se encontraron pares válidos).")

        master_df = pd.concat(master_rows, ignore_index=True)
        
        # 4. Standard Format
        std_df = _build_standard_from_master(master_df)
        
        # 5. Summary
        total_rows = len(std_df)
        total_offers = std_df["Total por renglón"].sum() if "Total por renglón" in std_df else 0.0
        
        summary = {
            "total_rows": total_rows,
            "total_offers": float(total_offers or 0.0),
            "awarded": 0.0,
            "pct_over_awarded": 0.0,
            "renglones": int(std_df["Renglón"].nunique()) if "Renglón" in std_df else 0
        }
        
        print(f"--- END NORMALIZE_LA_PAMPA ---\n")
        return {"df": std_df, "summary": summary}
        
    finally:
        # Cleanup extract dir
        try:
            shutil.rmtree(extract_dir)
        except Exception:
            pass

# ======== main ========
if __name__ == "__main__":
    run_all_pipelines()