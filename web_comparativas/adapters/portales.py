from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import unicodedata, re

BASE_DIR = Path(__file__).resolve().parents[1]  # .../web_comparativas
MARCAS_PATH = BASE_DIR / "marcas.xlsm"

def _norm(s):
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def _parse_latam(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("\xa0", " ").replace(" ", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _to_int(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return pd.NA
    try:
        return int(float(str(x).replace(",", ".").replace(" ", "")))
    except Exception:
        m = re.search(r"\d+", str(x))
        return int(m.group()) if m else pd.NA

def _cargar_marcas():
    if MARCAS_PATH.exists():
        try:
            df = pd.read_excel(MARCAS_PATH, header=None, usecols=[0])
            return [str(x).upper() for x in df[0].dropna().tolist()]
        except Exception:
            return []
    return []

MARCAS = _cargar_marcas()

def _detectar_marca(texto):
    if not texto:
        return "S/D"
    t = str(texto).upper()
    for m in MARCAS:
        if m in t:
            return m
    return "S/D"

# ============================================================================
# ALIASES DE COLUMNAS
# Mapeo de nombre canónico → lista de variantes normalizadas (sin tildes, lowercase).
# Usado por el parser de tabla plana y por _find_col_by_aliases.
# ============================================================================
COLUMN_ALIASES: dict[str, list[str]] = {
    "Renglón": [
        "renglon", "renglón", "nro renglon", "nro renglón", "nro. renglon",
        "n renglon", "n° renglon", "n renglon", "item", "ítem", "linea", "línea",
        "numero de renglon", "numero renglon",
    ],
    "Alternativa": ["alternativa", "alt.", "opcion", "opción", "variante"],
    "Código": [
        "codigo", "código", "cod.", "cod", "art.", "articulo", "artículo",
        "codigo de art", "codigo art",
    ],
    "Descripción": [
        "descripcion", "descripción", "detalle", "producto", "bien",
        "articulo", "artículo", "denominacion", "denominación", "objeto",
    ],
    "Cantidad solicitada": [
        "cantidad solicitada", "cant. solicitada", "cant solicitada",
        "cantidad pedida", "cantidad requerida", "cant. pedida", "cant pedida",
        "cantidad licitada",
    ],
    "Unidad de medida": [
        "unidad de medida", "u. medida", "unidad", "unid.", "um", "u/m",
    ],
    "Proveedor": [
        "proveedor", "oferente", "razon social", "razón social",
        "empresa", "firma", "adjudicatario", "licitante", "nombre proveedor",
    ],
    "Precio unitario": [
        "precio unitario", "precio unit.", "precio unit", "p. unitario",
        "p unitario", "valor unitario", "precio ofertado", "precio oferta",
        "precio neto", "pu", "imp. unitario", "importe unitario",
        "precio ofrecido", "pu ofertado",
    ],
    "Cantidad ofertada": [
        "cantidad ofertada", "cant. ofertada", "cant ofertada",
        "cantidad", "cant.",
    ],
    "Total por renglón": [
        "total por renglon", "total por renglón", "total renglon",
        "total renglón", "importe total", "monto total", "importe", "total",
        "monto", "total ofertado",
    ],
    "Especificación técnica": [
        "especificacion tecnica", "especificación técnica",
        "especificacion", "especificación", "detalle tecnico", "detalle técnico",
    ],
    "Observaciones": [
        "obs.", "obs", "observacion", "observación", "observaciones",
    ],
    "Marca": [
        "marca", "laboratorio", "fabricante", "lab.", "marca comercial",
    ],
}

# Columnas obligatorias para aceptar una tabla plana como válida
_FLAT_REQUIRED: set[str] = {"Proveedor", "Precio unitario"}
# Al menos una columna identificadora de ítem debe existir
_FLAT_IDENTIFIERS: set[str] = {"Renglón", "Descripción"}


def _find_col_by_aliases(canonical: str, normalized_headers: list) -> int | None:
    """
    Devuelve el índice de la primera columna que coincide con algún alias
    del nombre canónico, o None si no hay coincidencia.
    La comparación es por inclusión normalizada (sin tildes, lowercase).
    """
    for alias in COLUMN_ALIASES.get(canonical, []):
        tok = _norm(alias)
        for i, h in enumerate(normalized_headers):
            if h and (h == tok or tok in h or h in tok):
                return i
    return None


def _looks_like_flat_provider_table(row_headers) -> bool:
    """
    True si la fila de cabecera contiene columnas explícitas de
    Proveedor + Precio unitario + (Renglón o Descripción).
    Indica que el archivo es una tabla plana, no de bloques de proveedor.
    """
    hdrs_norm = [_norm(v) for v in row_headers.tolist()]
    has_prov  = _find_col_by_aliases("Proveedor",      hdrs_norm) is not None
    has_price = _find_col_by_aliases("Precio unitario", hdrs_norm) is not None
    has_id    = any(
        _find_col_by_aliases(c, hdrs_norm) is not None
        for c in _FLAT_IDENTIFIERS
    )
    return has_prov and has_price and has_id


# Tokens que identifican la fila de cabecera de proveedores en el comparativo.
# El primero con el mayor score gana. Se normalizan antes de comparar.
_PRICE_TOKENS = [
    "precio unitario",
    "precio unit",
    "p. unitario",
    "imp. unitario",
    "importe unitario",
    "pu ofertado",
    "precio ofertado",
    "precio oferta",
    "precio neto",
    "valor unitario",
]

def _find_header_and_provider_rows(df, token="precio unitario", search_rows=(1, 60)):
    """
    Busca la fila que contiene la columna de precio unitario (u equivalente).
    Retorna (header_row_idx, provider_row_idx) donde provider_row es la fila
    inmediatamente superior (con nombres de proveedores).

    Cambios respecto a versión anterior:
    - Rango ampliado: (1, 60) en vez de (5, 40) para tolerar archivos con
      filas de título/logo antes de la tabla.
    - Multi-token: si el token exacto no aparece, prueba variantes de la lista
      _PRICE_TOKENS para ser tolerante con nombres de columna diferentes.
    """
    all_tokens = [_norm(token)] + [_norm(t) for t in _PRICE_TOKENS if _norm(t) != _norm(token)]

    for r in range(search_rows[0], min(search_rows[1], df.shape[0])):
        row_vals = [_norm(v) for v in df.iloc[r].tolist()]
        for tok in all_tokens:
            if any(v and tok in v for v in row_vals):
                return r, r - 1
    return None, None

def _first_fixed_cols(row_headers):
    fixed = {}
    for j, v in enumerate(row_headers.tolist()):
        t = _norm(v)
        if not t:
            continue
        if "renglon" in t and "Renglón" not in fixed:
            fixed["Renglón"] = j
        elif ("alternativa" in t or "opcion" in t) and "Alternativa" not in fixed:
            fixed["Alternativa"] = j
        elif "codigo" in t and "Código" not in fixed:
            fixed["Código"] = j
        elif ("descripcion" in t or "detalle" in t or "producto" in t) and "Descripción" not in fixed:
            fixed["Descripción"] = j
        elif ("cantidad" in t or "cant." in t) and ("solic" in t or "pedid" in t or ("ofert" not in t and "adjud" not in t)) and "Cantidad solicitada" not in fixed:
            fixed["Cantidad solicitada"] = j
        elif ("unidad" in t or "u. medida" in t or "unid." in t) and "Unidad de medida" not in fixed:
            fixed["Unidad de medida"] = j
    return fixed

def _is_price_header(h: str) -> bool:
    """True si el encabezado normalizado corresponde a alguna variante de precio unitario."""
    return any(tok in h for tok in [_norm(t) for t in _PRICE_TOKENS])

def _provider_blocks(df, provider_row_idx, header_row_idx, fixed_cols):
    last_fixed = max(fixed_cols.values()) if fixed_cols else 0
    providers = []
    for j, v in enumerate(df.iloc[provider_row_idx].tolist()):
        if pd.isna(v):
            continue
        name = str(v).strip()
        if not name or _norm(name) in ("", "nan"):
            continue
        if j > last_fixed:
            providers.append((j, name))
    blocks = []
    for idx, (start, name) in enumerate(providers):
        end = providers[idx + 1][0] if idx + 1 < len(providers) else df.shape[1]
        hdrs_norm = [_norm(x) for x in df.iloc[header_row_idx, start:end].tolist()]
        if any(h and _is_price_header(h) for h in hdrs_norm):
            blocks.append((name, start, end))
    return blocks

def _map_cols_in_block(df, header_row_idx, start, end):
    hdrs = [_norm(v) for v in df.iloc[header_row_idx, start:end].tolist()]
    cols = list(range(start, end))
    mapping = {}
    for h, c in zip(hdrs, cols):
        if not h:
            continue
        # Precio unitario (usa la misma lista de tokens que la detección)
        if _is_price_header(h) and "total" not in h:
            mapping.setdefault("Precio unitario", c)

        # Cantidad ofertada
        if any(x in h for x in ("cantidad ofertada", "cant. ofertada", "cantidad", "cant.")):
            if "solicitada" not in h:
                mapping.setdefault("Cantidad ofertada", c)

        # Total por renglón
        if any(x in h for x in ("total por rengl", "importe total", "monto total", "total renglon")) or h.startswith("total"):
            mapping.setdefault("Total por renglón", c)

        # Especificación técnica / Marca
        if any(x in h for x in ("especificacion", "observacion", "marca", "detalle", "laboratorio", "fabricante")):
            mapping.setdefault("Especificación técnica", c)
    return mapping

def _build_summary_and_finalize(out: pd.DataFrame, diag: dict) -> tuple[pd.DataFrame, dict]:
    """
    Post-procesamiento común para ambos parsers (bloques y tabla plana):
    - Conversión de enteros (Renglón, Alternativa)
    - Cálculo de Total por renglón si falta
    - Cálculo de Posicion
    - KPIs (total_offers, awarded, pct_over_awarded)
    - Orden de columnas canónico
    Devuelve (df_final, summary_dict).
    """
    # Enteros
    for col in ("Renglón", "Alternativa"):
        if col in out.columns:
            out[col] = out[col].apply(_to_int).astype("Int64")

    # Total por renglón calculado si falta
    if "Total por renglón" not in out.columns or out["Total por renglón"].isna().all():
        pu   = pd.to_numeric(out.get("Precio unitario"),   errors="coerce").fillna(0)
        cant = pd.to_numeric(out.get("Cantidad ofertada"), errors="coerce").fillna(0)
        out["Total por renglón"] = pu * cant

    # Posición por renglón (rank de precio ascendente dentro del mismo renglón)
    group_col = next(
        (c for c in ("Renglón", "Descripción", "Código") if c in out.columns), None
    )
    if group_col and "Precio unitario" in out.columns:
        out["Posicion"] = (
            out.groupby([group_col])["Precio unitario"]
               .rank(method="dense", ascending=True)
               .astype("Int64")
        )

    # KPIs
    total_offers = float(pd.to_numeric(out["Total por renglón"], errors="coerce").fillna(0).sum())
    awarded = 0.0
    if "Posicion" in out.columns:
        winners = out[out["Posicion"] == 1]
        awarded = float(pd.to_numeric(winners["Total por renglón"], errors="coerce").fillna(0).sum())
    pct = round(awarded * 100.0 / total_offers, 2) if total_offers > 0 else 0.0

    # Charts (best-effort)
    try:
        chart_suppliers = (
            out.groupby("Proveedor")["Total por renglón"]
               .sum().sort_values(ascending=False).head(10)
               .pipe(lambda s: {"labels": s.index.tolist(), "values": [float(x) for x in s.values]})
        )
    except Exception:
        chart_suppliers = {"labels": [], "values": []}

    try:
        chart_positions = (
            out["Posicion"].value_counts(dropna=True).sort_index()
               .pipe(lambda s: {"labels": [str(i) for i in s.index.tolist()], "values": s.values.tolist()})
        ) if "Posicion" in out.columns else {"labels": [], "values": []}
    except Exception:
        chart_positions = {"labels": [], "values": []}

    summary = {
        "total_offers":    round(total_offers, 2),
        "awarded":         round(awarded, 2),
        "pct_over_awarded": pct,
        "chart_suppliers": chart_suppliers,
        "chart_positions": chart_positions,
        "__diag__": diag,
    }

    # Orden canónico de columnas (el tablero espera este orden)
    ordered = [
        "Proveedor", "Renglón", "Alternativa", "Código", "Descripción",
        "Cantidad solicitada", "Unidad de medida",
        "Precio unitario", "Cantidad ofertada", "Total por renglón",
        "Especificación técnica", "Observaciones", "Marca", "Posicion",
    ]
    cols = [c for c in ordered if c in out.columns] + [c for c in out.columns if c not in ordered]
    return out[cols], summary


def _parse_flat_provider_table(
    df: pd.DataFrame,
    header_row_idx: int,
    selected_sheet: str,
    all_sheets: list,
    detected_cols: list,
) -> tuple[pd.DataFrame, dict]:
    """
    Parser para archivos de comparativa en formato tabla plana, donde el
    proveedor es una columna explícita (en vez de un bloque de columnas).

    Ejemplo de cabecera detectada:
      Renglón | Alternativa | Precio unitario | Proveedor | Cantidad Ofertada | ...

    Devuelve (df_normalizado, diag_dict).
    Si falla, devuelve (DataFrame vacío, diag_dict con failed_stage).
    """
    import logging as _log
    _logger = _log.getLogger("web_comp.portales")

    row_headers = df.iloc[header_row_idx]
    hdrs_norm   = [_norm(v) for v in row_headers.tolist()]

    # Mapear cada nombre canónico al índice de columna real
    col_map: dict[str, int] = {}
    mapped_display: dict[str, str] = {}
    for canonical in COLUMN_ALIASES:
        idx = _find_col_by_aliases(canonical, hdrs_norm)
        if idx is not None:
            col_map[canonical] = idx
            raw_hdr = row_headers.iloc[idx] if idx < len(row_headers) else canonical
            mapped_display[canonical] = str(raw_hdr).strip()

    missing_required = _FLAT_REQUIRED - set(col_map.keys())
    has_identifier   = bool(_FLAT_IDENTIFIERS & set(col_map.keys()))

    diag: dict = {
        "parser_mode":         "flat_provider_table",
        "sheets_available":    all_sheets,
        "sheet_selected":      selected_sheet,
        "header_row_detected": header_row_idx,
        "columns_detected":    detected_cols,
        "mapped_columns":      mapped_display,
        "missing_required":    sorted(missing_required),
        "rows_read":           0,
        "rows_normalized":     0,
        "failed_stage":        None,
    }

    # Validar columnas obligatorias
    if missing_required or not has_identifier:
        missing_list = sorted(missing_required)
        if not has_identifier:
            missing_list.append("Renglón o Descripción")
        diag["failed_stage"] = "missing_required_columns"
        _logger.warning(
            "[portales/flat] Tabla plana detectada pero faltan columnas obligatorias: %s",
            missing_list,
        )
        return pd.DataFrame(), diag

    # Construir filas desde header_row_idx + 1
    data_start = header_row_idx + 1
    diag["rows_read"] = max(0, df.shape[0] - data_start)
    out_rows: list[dict] = []

    prov_idx  = col_map.get("Proveedor")
    price_idx = col_map.get("Precio unitario")

    for r in range(data_start, df.shape[0]):
        row = df.iloc[r]

        # Leer proveedor y precio para filtro de fila vacía
        prov_raw  = row[prov_idx]  if prov_idx  is not None and prov_idx  < len(row) else None
        price_raw = row[price_idx] if price_idx is not None and price_idx < len(row) else None

        prov_str  = str(prov_raw).strip()  if prov_raw  is not None and not (isinstance(prov_raw,  float) and pd.isna(prov_raw))  else ""
        price_val = _parse_latam(price_raw)

        # Saltar filas completamente vacías de datos útiles
        if not prov_str and price_val is None:
            continue

        rec: dict = {}
        for canonical, col_idx in col_map.items():
            v = row[col_idx] if col_idx < len(row) else None
            rec[canonical] = None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

        # Normalizar numéricos
        rec["Precio unitario"]    = _parse_latam(rec.get("Precio unitario"))
        rec["Cantidad ofertada"]  = _parse_latam(rec.get("Cantidad ofertada"))
        rec["Total por renglón"]  = _parse_latam(rec.get("Total por renglón"))
        if "Cantidad solicitada" in rec:
            rec["Cantidad solicitada"] = _parse_latam(rec.get("Cantidad solicitada"))

        # Proveedor como string limpio
        if rec.get("Proveedor") is not None:
            rec["Proveedor"] = str(rec["Proveedor"]).strip()

        # Detección de marca si no tiene columna propia
        if "Marca" not in col_map:
            espec = rec.get("Especificación técnica") or rec.get("Observaciones")
            rec["Marca"] = _detectar_marca(espec)

        out_rows.append(rec)

    diag["rows_normalized"] = len(out_rows)

    if not out_rows:
        diag["failed_stage"] = "no_valid_rows"
        _logger.warning(
            "[portales/flat] Cabecera detectada en '%s' fila %d, "
            "pero no se encontraron filas con Proveedor y Precio unitario.",
            selected_sheet, header_row_idx,
        )
        return pd.DataFrame(), diag

    out = pd.DataFrame(out_rows)

    _logger.info(
        "[portales/flat] Tabla plana procesada | filas=%d | proveedores únicos=%d | hoja='%s'",
        len(out),
        out["Proveedor"].nunique() if "Proveedor" in out.columns else 0,
        selected_sheet,
    )

    return out, diag


def _normalize_core(path: Path, hoja: str | None = None):
    import logging as _log
    _logger = _log.getLogger("web_comp.portales")

    wb = load_workbook(path, data_only=True)
    all_sheets = wb.sheetnames
    _logger.info("[portales] Hojas disponibles en %s: %s", path.name, all_sheets)

    # Selección de hoja: preferir la indicada, luego buscar la más prometedora
    def _sheet_score(ws_name: str) -> int:
        n = ws_name.lower()
        score = 0
        for keyword in ("comparativo", "cuadro", "oferta", "licitacion", "proveedores", "precios"):
            if keyword in n:
                score += 2
        ws_candidate = wb[ws_name]
        # Premiar hojas con más datos
        if ws_candidate.max_row and ws_candidate.max_row > 5:
            score += 1
        return score

    if hoja and hoja in all_sheets:
        ws = wb[hoja]
        selected_sheet = hoja
    else:
        # Elegir la hoja con mayor score; como tiebreak, la primera
        best = max(all_sheets, key=_sheet_score)
        ws = wb[best]
        selected_sheet = best
        if hoja:
            _logger.warning("[portales] Hoja '%s' no encontrada, usando '%s'", hoja, selected_sheet)

    _logger.info("[portales] Hoja seleccionada: '%s'", selected_sheet)

    data = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    df = pd.DataFrame(data)

    header_row_idx, provider_row_idx = _find_header_and_provider_rows(df)

    _diag_empty = {
        "total_offers": 0.0, "awarded": 0.0, "pct_over_awarded": 0.0,
        "__diag__": {
            "sheets_available": all_sheets,
            "sheet_selected": selected_sheet,
            "header_row_detected": None,
            "columns_detected": [],
            "blocks_detected": 0,
            "failed_stage": None,
        }
    }

    if header_row_idx is None or provider_row_idx is None:
        # Intentar con TODAS las demás hojas antes de rendirse
        for alt_sheet in all_sheets:
            if alt_sheet == selected_sheet:
                continue
            ws2 = wb[alt_sheet]
            data2 = [[cell for cell in row] for row in ws2.iter_rows(values_only=True)]
            df2 = pd.DataFrame(data2)
            h2, p2 = _find_header_and_provider_rows(df2)
            if h2 is not None:
                _logger.info("[portales] Cabecera encontrada en hoja alternativa '%s' fila %d", alt_sheet, h2)
                df, header_row_idx, provider_row_idx, selected_sheet = df2, h2, p2, alt_sheet
                break

        if header_row_idx is None:
            _diag_empty["__diag__"]["failed_stage"] = "header_not_found"
            _logger.warning(
                "[portales] No se encontró fila de cabecera con 'precio unitario' (u equivalente) "
                "en ninguna hoja de %s. Filas analizadas: %d", path.name, df.shape[0]
            )
            return pd.DataFrame(), _diag_empty

    row_headers = df.iloc[header_row_idx]
    detected_cols = [str(v).strip() for v in row_headers.tolist() if v and str(v).strip()]
    _logger.info("[portales] Fila de cabecera: %d | Columnas detectadas: %s", header_row_idx, detected_cols[:20])

    fixed_cols = _first_fixed_cols(row_headers)
    blocks = _provider_blocks(df, provider_row_idx, header_row_idx, fixed_cols)

    _diag_empty["__diag__"].update({
        "sheet_selected": selected_sheet,
        "header_row_detected": header_row_idx,
        "columns_detected": detected_cols,
        "blocks_detected": len(blocks),
    })

    # ── Fallo por falta de columnas fijas (Renglón, etc.) ──────────────────────
    if not fixed_cols:
        _diag_empty["__diag__"]["failed_stage"] = "no_fixed_cols"
        _logger.warning(
            "[portales] No se detectaron columnas fijas (Renglón, Descripción, etc.) en '%s'.",
            selected_sheet,
        )
        return pd.DataFrame(), _diag_empty

    # ── Sin bloques de proveedor: intentar parser de tabla plana ────────────
    if not blocks:
        _logger.info(
            "[portales] No hay bloques de proveedor en '%s'. "
            "Verificando si es tabla plana (Proveedor como columna)...",
            selected_sheet,
        )
        if _looks_like_flat_provider_table(row_headers):
            flat_out, flat_diag = _parse_flat_provider_table(
                df, header_row_idx, selected_sheet, all_sheets, detected_cols
            )
            if flat_out is not None and not flat_out.empty:
                _logger.info(
                    "[portales] Tabla plana procesada correctamente (%d filas).", len(flat_out)
                )
                flat_diag.update({"failed_stage": None})
                return _build_summary_and_finalize(flat_out, flat_diag)
            else:
                # El flat parser falló también; devolvemos su diagnóstico
                _diag_empty["__diag__"].update(flat_diag)
                return pd.DataFrame(), _diag_empty
        else:
            _diag_empty["__diag__"]["failed_stage"] = "no_provider_blocks"
            _logger.warning(
                "[portales] Sin bloques de proveedor y sin columna 'Proveedor' explícita. "
                "fixed_cols=%s", list(fixed_cols.keys()),
            )
            return pd.DataFrame(), _diag_empty

    # ── Parser de bloques (camino original) ─────────────────────────────────
    out_rows = []
    for r in range(header_row_idx + 1, df.shape[0]):
        row = df.iloc[r]
        rc = fixed_cols.get("Renglón", None)
        if rc is None or pd.isna(row[rc]):
            continue

        base = {}
        for k, i in fixed_cols.items():
            base[k] = row[i] if i < len(row) else None

        for (prov_name, start, end) in blocks:
            mapping = _map_cols_in_block(df, header_row_idx, start, end)
            rec = base.copy()
            rec["Proveedor"] = str(prov_name).strip()

            pu    = _parse_latam(row[mapping["Precio unitario"]])    if "Precio unitario"    in mapping else None
            co    = _parse_latam(row[mapping["Cantidad ofertada"]])  if "Cantidad ofertada"  in mapping else None
            tot   = _parse_latam(row[mapping["Total por renglón"]])  if "Total por renglón"  in mapping else None
            espec = row[mapping["Especificación técnica"]]           if "Especificación técnica" in mapping else None

            rec["Precio unitario"]     = pu
            rec["Cantidad ofertada"]   = co
            rec["Total por renglón"]   = tot
            rec["Especificación técnica"] = espec
            rec["Marca"] = _detectar_marca(espec)
            out_rows.append(rec)

    _logger.info(
        "[portales] Filas extraídas: %d | Bloques de proveedor: %d | Hoja: '%s'",
        len(out_rows), len(blocks), selected_sheet,
    )

    if not out_rows:
        _diag_empty["__diag__"].update({"failed_stage": "no_data_rows"})
        _logger.warning("[portales] Bloques de proveedor detectados pero sin filas de datos.")
        return pd.DataFrame(), _diag_empty

    block_diag = {
        "parser_mode":         "provider_blocks",
        "sheets_available":    all_sheets,
        "sheet_selected":      selected_sheet,
        "header_row_detected": header_row_idx,
        "columns_detected":    detected_cols,
        "blocks_detected":     len(blocks),
        "rows_extracted":      len(out_rows),
        "failed_stage":        None,
    }
    return _build_summary_and_finalize(pd.DataFrame(out_rows), block_diag)

# ====== INTERFACES PÚBLICAS ======

def normalize_comprar(input_path: Path, metadata: dict, out_dir: Path):
    df, summary = _normalize_core(input_path, hoja="Cuadro comparativo")
    return {"df": df, "summary": summary}

def normalize_bac(input_path: Path, metadata: dict, out_dir: Path):
    # De momento usamos la misma lógica; si el BAC cambia, clonás y ajustás.
    df, summary = _normalize_core(input_path, hoja=None)
    return {"df": df, "summary": summary}

def normalize_pbac(input_path: Path, metadata: dict, out_dir: Path):
    df, summary = _normalize_core(input_path, hoja=None)
    return {"df": df, "summary": summary}
