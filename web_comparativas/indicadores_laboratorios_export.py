"""
Exportación a Excel (.xlsx) de Informes de Laboratorio — Indicadores Comerciales.

Tercera salida del módulo, junto al PDF ("Imprimir") y al CSV de la barra de filtros;
NO reemplaza ni modifica a ninguna de las dos.

Fuente de datos: indicadores_laboratorios_service.get_export_bundle(), es decir la MISMA
consulta que alimenta la vista (tablas summary + corrida activa vía _corrida_activa(), con
fallback al vivo). Acá no hay SQL ni reglas de negocio nuevas: solo se arma el libro.

Hojas, en orden:
  1. "Resumen"                      — período, filtros, KPIs de las tarjetas,
                                      Evolución Mensual de Unidades y ranking Top Laboratorios.
  2. "Sell Out por Producto"        — matriz producto × meses (colapsa clientes).
  3. "Sell Out por Marca y Cliente" — matriz cliente + marca × meses, orden de la vista
                                      (unidades desc).
  4. "Sell Out por Cliente y Marca" — misma matriz ordenada por cliente y luego marca.

Las tres matrices llevan TODAS las filas del resultado filtrado (la pantalla recorta a
1500 por scroll y /detalle a 2000; acá no se recorta).
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from web_comparativas.indicadores_laboratorios_service import get_export_bundle, normalize_text

logger = logging.getLogger("wc.indicadores.lab.export")

NUM_FMT = "#,##0"
PCT_FMT = "+0.0%;-0.0%"

_MESES_ABBR = ("ene", "feb", "mar", "abr", "may", "jun",
               "jul", "ago", "sep", "oct", "nov", "dic")

_FONT_TITLE = Font(bold=True, size=14)
_FONT_BOLD = Font(bold=True)
_FONT_MUTED = Font(size=9, color="FF6B7280")
_FILL_HEAD = PatternFill("solid", fgColor="FFEFF3F8")
_ALIGN_RIGHT = Alignment(horizontal="right")
_ALIGN_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ─── Helpers de formato ──────────────────────────────────────────────────────

def _fmt_month(mes: str) -> str:
    """'2025-06' → 'jun-25' (mismo label que la grilla en pantalla)."""
    try:
        year, month = mes.split("-")[:2]
        return f"{_MESES_ABBR[int(month) - 1]}-{year[2:]}"
    except (ValueError, IndexError, AttributeError):
        return str(mes or "")


def _fmt_dmy(value: Optional[date]) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def _sort_key(value) -> str:
    """Orden alfabético insensible a acentos/mayúsculas (equivalente al
    localeCompare('es') que usa la vista)."""
    return normalize_text(value or "")


# ─── Reordenamientos que hoy hace el front (mismos datos, otra vista) ────────

def _build_product_rows(detalle: list) -> list:
    """Agrega el detalle por marca sumando unidades y meses entre clientes.
    Espejo de _buildProductDetail() en indicadores_laboratorios.js."""
    grouped: dict = {}
    for row in detalle:
        key = row.get("marca") or "SIN PRODUCTO"
        item = grouped.setdefault(key, {"marca": key, "unidades": 0.0, "mensual": {}})
        item["unidades"] += float(row.get("unidades") or 0)
        for mes, value in (row.get("mensual") or {}).items():
            item["mensual"][mes] = item["mensual"].get(mes, 0.0) + float(value or 0)
    return sorted(grouped.values(), key=lambda r: r["unidades"], reverse=True)


def _sort_by_cliente_marca(detalle: list) -> list:
    """Espejo de _sortRows(_detail, 'cliente', 'marca'): cliente, luego marca,
    luego unidades desc."""
    return sorted(
        detalle,
        key=lambda r: (_sort_key(r.get("cliente")), _sort_key(r.get("marca")),
                       -float(r.get("unidades") or 0)),
    )


# ─── Escritura de hojas ──────────────────────────────────────────────────────

def _write_matrix_sheet(wb: Workbook, title: str, fixed_cols: list, rows: list, months: list):
    """Una hoja de matriz: columnas fijas + una columna por mes + TOTAL.

    fixed_cols: [(label, getter, ancho)]. Encabezado en negrita y congelado junto a
    las columnas fijas. Celda VACÍA cuando no hay dato (igual que la vista, que tampoco
    muestra ceros); los números van como número real con formato de miles.
    """
    ws = wb.create_sheet(title=title)

    header = [label for label, _getter, _width in fixed_cols]
    header += [_fmt_month(m) for m in months]
    header.append("TOTAL")
    ws.append(header)
    for cell in ws[1]:
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEAD
        cell.alignment = _ALIGN_HEAD

    n_fixed = len(fixed_cols)
    for row in rows:
        mensual = row.get("mensual") or {}
        values = [getter(row) for _label, getter, _width in fixed_cols]
        for mes in months:
            v = mensual.get(mes)
            # Sin dato (o 0) → celda vacía, como en pantalla.
            values.append(None if v in (None, 0) else float(v))
        values.append(float(row.get("unidades") or 0))
        ws.append(values)

    last_row = ws.max_row
    if last_row > 1:
        for col in range(n_fixed + 1, n_fixed + len(months) + 2):
            letter = get_column_letter(col)
            for cell in ws[f"{letter}2":f"{letter}{last_row}"]:
                cell[0].number_format = NUM_FMT
                cell[0].alignment = _ALIGN_RIGHT
        total_letter = get_column_letter(n_fixed + len(months) + 1)
        for cell in ws[f"{total_letter}2":f"{total_letter}{last_row}"]:
            cell[0].font = _FONT_BOLD

    for idx, (_label, _getter, width) in enumerate(fixed_cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx in range(n_fixed + 1, n_fixed + len(months) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 12
    ws.column_dimensions[get_column_letter(n_fixed + len(months) + 1)].width = 14

    ws.freeze_panes = f"{get_column_letter(n_fixed + 1)}2"
    return ws


def _write_resumen_sheet(wb: Workbook, resumen: dict, meta: dict):
    ws = wb.create_sheet(title="Resumen")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20

    ws["A1"] = "Informes de Laboratorio"
    ws["A1"].font = _FONT_TITLE
    ws["A2"] = "Suite SIEM · Indicadores Comerciales — Sell out mensual de unidades"
    ws["A2"].font = _FONT_MUTED

    ws["A4"] = "Período"
    ws["B4"] = f"{_fmt_dmy(meta['desde'])} — {_fmt_dmy(meta['hasta_display'])}"
    ws["A5"] = "Filtros"
    ws["B5"] = meta["filtros"] or "Sin filtros adicionales"
    ws["A6"] = "Emitido"
    ws["B6"] = meta["emitido"]
    for r in (4, 5, 6):
        ws[f"A{r}"].font = _FONT_BOLD

    meses = resumen.get("meses") or []
    labs = resumen.get("laboratorios") or []
    marcas = resumen.get("marcas") or []
    top_lab = labs[0] if labs else None
    top_marca = marcas[0] if marcas else None

    row = 8
    ws[f"A{row}"] = "Indicadores"
    ws[f"A{row}"].font = _FONT_BOLD
    row += 1
    ws.cell(row=row, column=1, value="Indicador").font = _FONT_BOLD
    ws.cell(row=row, column=2, value="Valor").font = _FONT_BOLD
    ws.cell(row=row, column=3, value="Detalle").font = _FONT_BOLD
    for cell in (ws.cell(row=row, column=1), ws.cell(row=row, column=2), ws.cell(row=row, column=3)):
        cell.fill = _FILL_HEAD
    row += 1

    variacion = resumen.get("variacion_mensual")
    kpis = [
        ("Unidades Totales", float(resumen.get("total_unidades") or 0), NUM_FMT, "período seleccionado"),
        ("Promedio Mensual", float(resumen.get("promedio_mensual") or 0), NUM_FMT, "unidades por mes activo"),
        ("Variación Mensual", None if variacion is None else float(variacion), PCT_FMT, "último mes vs anterior"),
        ("Mayor Laboratorio", top_lab["name"] if top_lab else "—", None,
         f"{round(top_lab['value']):,.0f} unidades".replace(",", ".") if top_lab else "—"),
        ("Laboratorios", float(resumen.get("cantidad_laboratorios") or 0), NUM_FMT, "con ventas en el filtro"),
        ("Marcas Comerciales", float(resumen.get("cantidad_marcas") or 0), NUM_FMT, "productos activos"),
        ("Clientes", float(resumen.get("cantidad_clientes") or 0), NUM_FMT, "grupos con unidades"),
        ("Marca Líder", top_marca["name"] if top_marca else "—", None,
         f"{round(top_marca['value']):,.0f} unidades".replace(",", ".") if top_marca else "—"),
    ]
    for label, value, fmt, detail in kpis:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value="—" if value is None else value)
        if fmt and value is not None:
            cell.number_format = fmt
            cell.alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=3, value=detail).font = _FONT_MUTED
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Evolución Mensual de Unidades").font = _FONT_BOLD
    row += 1
    for col, label in ((1, "Mes"), (2, "Unidades")):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEAD
    row += 1
    for item in meses:
        ws.cell(row=row, column=1, value=_fmt_month(item.get("mes")))
        cell = ws.cell(row=row, column=2, value=float(item.get("unidades") or 0))
        cell.number_format = NUM_FMT
        cell.alignment = _ALIGN_RIGHT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top Laboratorios").font = _FONT_BOLD
    row += 1
    for col, label in ((1, "Laboratorio"), (2, "Unidades")):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEAD
    row += 1
    for item in labs:
        ws.cell(row=row, column=1, value=item.get("name"))
        cell = ws.cell(row=row, column=2, value=float(item.get("value") or 0))
        cell.number_format = NUM_FMT
        cell.alignment = _ALIGN_RIGHT
        row += 1

    return ws


# ─── Entrada pública ─────────────────────────────────────────────────────────

def build_laboratorios_workbook(
    desde: date,
    hasta: date,
    laboratorio: Optional[str] = None,
    familia: Optional[str] = None,
    cliente: Optional[str] = None,
    search: Optional[str] = None,
    cadneg: Optional[str] = None,
) -> tuple:
    """Devuelve (bytes_del_xlsx, nombre_de_archivo).

    `hasta` llega EXCLUSIVO (convención del módulo: fecha < hasta; el front manda
    hasta+1 día), así que la fecha que ve el usuario —y la que va al nombre del
    archivo— es hasta-1.
    """
    bundle = get_export_bundle(desde, hasta, laboratorio=laboratorio, familia=familia,
                               cliente=cliente, search=search, cadneg=cadneg)
    resumen = bundle["resumen"]
    detalle = bundle["detalle"]
    months = [m["mes"] for m in (resumen.get("meses") or [])]

    hasta_display = hasta - timedelta(days=1)
    filtros = "  ·  ".join(part for part in (
        f"Laboratorio: {laboratorio}" if laboratorio else "",
        f"Familia: {familia}" if familia else "",
        f"Cliente: {cliente}" if cliente else "",
        f"Negocio: {cadneg}" if cadneg else "",
        f"Búsqueda: {search}" if search else "",
    ) if part)

    wb = Workbook()
    wb.remove(wb.active)

    _write_resumen_sheet(wb, resumen, {
        "desde": desde,
        "hasta_display": hasta_display,
        "filtros": filtros,
        "emitido": datetime.now().strftime("%d/%m/%Y %H:%M"),
    })

    _write_matrix_sheet(
        wb, "Sell Out por Producto",
        [("Producto", lambda r: r.get("marca") or "—", 46)],
        _build_product_rows(detalle), months,
    )
    _write_matrix_sheet(
        wb, "Sell Out por Marca y Cliente",
        [("Cliente", lambda r: r.get("cliente") or "—", 42),
         ("Marca Comercial", lambda r: r.get("marca") or "—", 40)],
        detalle, months,
    )
    _write_matrix_sheet(
        wb, "Sell Out por Cliente y Marca",
        [("Cliente", lambda r: r.get("cliente") or "—", 42),
         ("Marca Comercial", lambda r: r.get("marca") or "—", 40)],
        _sort_by_cliente_marca(detalle), months,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (f"Informes_Laboratorio_{desde.strftime('%Y%m%d')}"
                f"_{hasta_display.strftime('%Y%m%d')}.xlsx")
    logger.info("lab export xlsx: filas detalle=%d meses=%d archivo=%s",
                len(detalle), len(months), filename)
    return buf.getvalue(), filename
